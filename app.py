from io import BytesIO
import pandas as pd
from datetime import datetime
import streamlit as st
from openpyxl.utils import get_column_letter
from pathlib import Path
import zipfile

from database import BACKUPS_DIR, DB_NAME, DOCUMENTOS_DIR, EVIDENCIAS_DIR, crear_tablas
from google_sheets import (
    actualizar_campos_por_clave,
    buscar_fila_por_clave,
    construir_tabla_resumen_pacientes,
    construir_tabla_tomas_pendientes,
    leer_sheet_como_dataframe,
    obtener_columnas_por_toma,
    obtener_fila_como_diccionario,
    preparar_datos_hospitalarios,
    preparar_resumen_influenza_observaciones,
)
from logic import (
    aprobar_usuario,
    aprobar_cierre_accion,
    aprobar_cierre_no_conformidad,
    aprobar_version_documento,
    actualizar_estado_accion_calidad,
    actualizar_accion_calidad,
    actualizar_estado_no_conformidad,
    actualizar_no_conformidad,
    autenticar_usuario,
    buscar_voluntario_por_id,
    contar_acciones_abiertas,
    contar_acciones_vencidas,
    contar_no_conformidades_abiertas,
    contar_racks_activos,
    contar_visitas_pendientes,
    contar_voluntarios,
    configurar_persistencia_calidad_supabase,
    configurar_persistencia_usuarios_supabase,
    descargar_documento_calidad,
    descargar_evidencia_calidad,
    exportar_a_excel,
    guardar_evidencia_calidad,
    listar_auditorias_calidad,
    listar_acciones_calidad,
    listar_bitacora_calidad,
    listar_documentos_calidad,
    listar_evidencias_calidad,
    listar_hallazgos_auditoria,
    listar_no_conformidades,
    listar_usuarios,
    listar_versiones_documento,
    obtener_ocupacion_racks,
    obtener_backend_calidad,
    obtener_backend_usuarios,
    obtener_usuarios_pendientes,
    registrar_auditoria_calidad,
    registrar_accion_calidad,
    registrar_documento_calidad,
    registrar_hallazgo_auditoria,
    registrar_no_conformidad,
    registrar_version_documento,
    registrar_usuario,
    registrar_voluntario,
    actualizar_rol_usuario,
    ver_alicuotas_pbmc_voluntario,
    ver_alicuotas_suero_voluntario,
    ver_rack_pbmc,
    ver_rack_suero,
    ver_racks,
    ver_visitas,
)
from logic import crear_admin_inicial

st.set_page_config(page_title="Sistema INER - Voluntarios", layout="wide")
crear_tablas()

configurar_persistencia_calidad_supabase(
    url=st.secrets.get("supabase_url", ""),
    key=st.secrets.get("supabase_key", ""),
    enabled=bool(st.secrets.get("use_supabase_quality", False)),
    evidencias_bucket=st.secrets.get("supabase_quality_evidencias_bucket", "calidad-evidencias"),
    documentos_bucket=st.secrets.get("supabase_quality_documentos_bucket", "calidad-documentos"),
)
configurar_persistencia_usuarios_supabase(
    url=st.secrets.get("supabase_url", ""),
    key=st.secrets.get("supabase_key", ""),
    enabled=bool(st.secrets.get("use_supabase_users", False)),
    table_name=st.secrets.get("supabase_users_table", "usuarios_app"),
)

crear_admin_inicial(
    st.secrets["admin_email"],
    st.secrets["admin_password"]
)

ROLES_USUARIO = ["captura", "responsable", "auditor", "calidad", "admin"]


def rol_actual():
    return st.session_state.get("rol_usuario", "captura")


def tiene_rol(*roles):
    rol = rol_actual()
    return rol == "admin" or rol in roles


def mostrar_aviso_permiso(mensaje):
    st.info(f"Permiso requerido: {mensaje}")


def crear_respaldo_sistema():
    marca_tiempo = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta_zip = BACKUPS_DIR / f"iner_respaldo_{marca_tiempo}.zip"

    with zipfile.ZipFile(ruta_zip, "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
        if DB_NAME.exists():
            zip_file.write(DB_NAME, arcname=DB_NAME.name)

        for carpeta in [EVIDENCIAS_DIR, DOCUMENTOS_DIR]:
            if carpeta.exists():
                for archivo in carpeta.rglob("*"):
                    if archivo.is_file():
                        zip_file.write(archivo, arcname=str(Path(carpeta.name) / archivo.relative_to(carpeta)))

    return ruta_zip


def listar_respaldos():
    if not BACKUPS_DIR.exists():
        return []
    return sorted(BACKUPS_DIR.glob("*.zip"), key=lambda ruta: ruta.stat().st_mtime, reverse=True)


def aplicar_estilos():
    st.markdown(
        """
        <style>
        :root {
            --iner-ink: #16324f;
            --iner-accent: #1f7a8c;
            --iner-soft: #e7f4f4;
            --iner-warm: #f3efe7;
            --iner-border: rgba(22, 50, 79, 0.12);
        }

        .block-container {
            padding-top: 2rem;
        }

        .iner-hero {
            background: linear-gradient(135deg, rgba(31,122,140,0.16), rgba(243,239,231,0.9));
            border: 1px solid var(--iner-border);
            border-radius: 22px;
            padding: 1.3rem 1.4rem;
            margin-bottom: 1rem;
        }

        .iner-kicker {
            color: var(--iner-accent);
            font-size: 0.86rem;
            font-weight: 700;
            letter-spacing: 0.08em;
            text-transform: uppercase;
            margin-bottom: 0.35rem;
        }

        .iner-title {
            color: var(--iner-ink);
            font-size: 2.1rem;
            line-height: 1.1;
            font-weight: 800;
            margin: 0;
        }

        .iner-copy {
            color: rgba(22, 50, 79, 0.82);
            margin-top: 0.55rem;
            margin-bottom: 0;
        }

        .iner-section {
            background: rgba(255, 255, 255, 0.78);
            border: 1px solid var(--iner-border);
            border-radius: 18px;
            padding: 0.9rem 1rem 0.4rem 1rem;
            margin: 0.7rem 0 1rem 0;
        }

        .iner-chip {
            display: inline-block;
            background: var(--iner-soft);
            color: var(--iner-ink);
            border-radius: 999px;
            padding: 0.25rem 0.75rem;
            font-size: 0.85rem;
            font-weight: 600;
            margin-bottom: 0.45rem;
        }

        div[data-testid="stMetric"] {
            background: linear-gradient(180deg, rgba(255,255,255,0.96), rgba(231,244,244,0.92));
            border: 1px solid var(--iner-border);
            border-radius: 18px;
            padding: 0.8rem;
        }

        div[data-testid="stDataFrame"] {
            border-radius: 16px;
            overflow: hidden;
        }

        div[data-testid="stDownloadButton"] button {
            background: linear-gradient(135deg, #1f7a8c, #16324f);
            color: white;
            border: none;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def tarjeta_seccion(etiqueta, titulo, descripcion=None):
    descripcion_html = f'<p class="iner-copy">{descripcion}</p>' if descripcion else ""
    st.markdown(
        f"""
        <div class="iner-section">
            <div class="iner-chip">{etiqueta}</div>
            <h3 style="margin:0;color:#16324f;">{titulo}</h3>
            {descripcion_html}
        </div>
        """,
        unsafe_allow_html=True,
    )


aplicar_estilos()


def semaforo_no_conformidad(estado, severidad):
    if estado == "Cerrada":
        return "Verde"
    if severidad in ["Crítica", "Alta"]:
        return "Rojo"
    if estado == "En proceso" or severidad == "Media":
        return "Amarillo"
    return "Verde"


def semaforo_accion(estado, fecha_compromiso):
    if estado == "Cerrada":
        return "Verde"

    fecha = pd.to_datetime(fecha_compromiso, errors="coerce")
    hoy = pd.Timestamp.today().normalize()

    if pd.notna(fecha) and fecha < hoy:
        return "Rojo"
    if estado == "En proceso":
        return "Amarillo"
    return "Verde"


def _color_semaforo(valor):
    colores = {
        "Rojo": "background-color: #f8d7da; color: #8a1c1c;",
        "Amarillo": "background-color: #fff3cd; color: #7a5d00;",
        "Verde": "background-color: #d1e7dd; color: #0f5132;",
    }
    return colores.get(valor, "")


def dataframe_con_semaforo(df):
    if df.empty or "Semáforo" not in df.columns:
        return df
    return df.style.map(_color_semaforo, subset=["Semáforo"])


def clasificar_vencimiento(fecha, dias_alerta=7):
    fecha_dt = pd.to_datetime(fecha, errors="coerce")
    if pd.isna(fecha_dt):
        return "Sin fecha"

    hoy = pd.Timestamp.today().normalize()
    delta = (fecha_dt.normalize() - hoy).days

    if delta < 0:
        return "Vencido"
    if delta == 0:
        return "Vence hoy"
    if delta <= dias_alerta:
        return "Vence pronto"
    return "En plazo"


def prioridad_desde_alerta(alerta):
    prioridades = {
        "Vencido": "Crítica",
        "Vence hoy": "Alta",
        "Vence pronto": "Media",
        "En plazo": "Baja",
        "Sin fecha": "Baja",
    }
    return prioridades.get(alerta, "Baja")


def construir_recordatorios_calidad(
    df_nc,
    df_acciones,
    df_auditorias,
    df_documentos,
    df_hallazgos=None,
    df_versiones=None,
):
    recordatorios = []

    if not df_nc.empty:
        for _, fila in df_nc.iterrows():
            estado_plazo = clasificar_vencimiento(fila.get("Fecha compromiso"), dias_alerta=7)
            if fila.get("Estado") != "Cerrada" and estado_plazo in ["Vencido", "Vence hoy", "Vence pronto"]:
                recordatorios.append(
                    {
                        "Tipo": "No conformidad",
                        "Referencia": fila.get("Código"),
                        "Título": fila.get("Título"),
                        "Responsable": fila.get("Responsable"),
                        "Fecha crítica": fila.get("Fecha compromiso"),
                        "Alerta": estado_plazo,
                        "Prioridad": prioridad_desde_alerta(estado_plazo),
                        "Motivo": f"No conformidad {str(fila.get('Severidad', '')).lower()} con compromiso próximo.",
                    }
                )

    if not df_acciones.empty:
        for _, fila in df_acciones.iterrows():
            estado_plazo = clasificar_vencimiento(fila.get("Fecha compromiso"), dias_alerta=7)
            if fila.get("Estado") != "Cerrada" and estado_plazo in ["Vencido", "Vence hoy", "Vence pronto"]:
                recordatorios.append(
                    {
                        "Tipo": "Acción",
                        "Referencia": fila.get("Código NC"),
                        "Título": fila.get("Título"),
                        "Responsable": fila.get("Responsable"),
                        "Fecha crítica": fila.get("Fecha compromiso"),
                        "Alerta": estado_plazo,
                        "Prioridad": prioridad_desde_alerta(estado_plazo),
                        "Motivo": f"Acción {str(fila.get('Tipo', '')).lower()} con fecha compromiso cercana.",
                    }
                )

    if not df_auditorias.empty:
        for _, fila in df_auditorias.iterrows():
            estado_plazo = clasificar_vencimiento(fila.get("Fecha programada"), dias_alerta=14)
            if fila.get("Estado") != "Cerrada" and estado_plazo in ["Vencido", "Vence hoy", "Vence pronto"]:
                recordatorios.append(
                    {
                        "Tipo": "Auditoría",
                        "Referencia": fila.get("Código"),
                        "Título": fila.get("Título"),
                        "Responsable": fila.get("Auditor líder"),
                        "Fecha crítica": fila.get("Fecha programada"),
                        "Alerta": estado_plazo,
                        "Prioridad": prioridad_desde_alerta(estado_plazo),
                        "Motivo": "Auditoría programada pendiente de ejecución o cierre.",
                    }
                )

    if not df_documentos.empty:
        for _, fila in df_documentos.iterrows():
            estado_plazo = clasificar_vencimiento(fila.get("Vigente hasta"), dias_alerta=30)
            if fila.get("Estado") == "Vigente" and estado_plazo in ["Vencido", "Vence hoy", "Vence pronto"]:
                recordatorios.append(
                    {
                        "Tipo": "Documento",
                        "Referencia": fila.get("Código"),
                        "Título": fila.get("Nombre"),
                        "Responsable": fila.get("Aprobado por"),
                        "Fecha crítica": fila.get("Vigente hasta"),
                        "Alerta": estado_plazo,
                        "Prioridad": prioridad_desde_alerta(estado_plazo),
                        "Motivo": "Documento vigente próximo a vencerse u obsoleto.",
                    }
                )

    if df_hallazgos is not None and not df_hallazgos.empty:
        for _, fila in df_hallazgos.iterrows():
            estado_plazo = clasificar_vencimiento(fila.get("Fecha compromiso"), dias_alerta=7)
            if fila.get("Estado") != "Cerrado" and estado_plazo in ["Vencido", "Vence hoy", "Vence pronto"]:
                recordatorios.append(
                    {
                        "Tipo": "Hallazgo",
                        "Referencia": fila.get("Referencia"),
                        "Título": fila.get("Descripción"),
                        "Responsable": fila.get("Responsable"),
                        "Fecha crítica": fila.get("Fecha compromiso"),
                        "Alerta": estado_plazo,
                        "Prioridad": prioridad_desde_alerta(estado_plazo),
                        "Motivo": f"Hallazgo de auditoría {str(fila.get('Severidad', '')).lower()} con seguimiento pendiente.",
                    }
                )

    if df_versiones is not None and not df_versiones.empty:
        versiones_pendientes = df_versiones[
            df_versiones["Aprobado por"].isna() | (df_versiones["Aprobado por"].astype(str).str.strip() == "")
        ]
        for _, fila in versiones_pendientes.iterrows():
            recordatorios.append(
                {
                    "Tipo": "Versión documental",
                    "Referencia": fila.get("Código"),
                    "Título": fila.get("Nombre"),
                    "Responsable": fila.get("Elaborado por"),
                    "Fecha crítica": fila.get("Registrada el"),
                    "Alerta": "Vence hoy",
                    "Prioridad": "Alta",
                    "Motivo": f"La versión {fila.get('Versión')} sigue pendiente de aprobación formal.",
                }
            )

    if not recordatorios:
        return pd.DataFrame(
            columns=["Tipo", "Referencia", "Título", "Responsable", "Fecha crítica", "Alerta", "Prioridad", "Motivo"]
        )

    df_recordatorios = pd.DataFrame(recordatorios)
    prioridad_alerta = {"Vencido": 0, "Vence hoy": 1, "Vence pronto": 2, "En plazo": 3, "Sin fecha": 4}
    prioridad_negocio = {"Crítica": 0, "Alta": 1, "Media": 2, "Baja": 3}
    df_recordatorios["orden_alerta"] = df_recordatorios["Alerta"].map(prioridad_alerta).fillna(99)
    df_recordatorios["orden_prioridad"] = df_recordatorios["Prioridad"].map(prioridad_negocio).fillna(99)
    df_recordatorios["Fecha_sort"] = pd.to_datetime(df_recordatorios["Fecha crítica"], errors="coerce")
    df_recordatorios = df_recordatorios.sort_values(
        by=["orden_prioridad", "orden_alerta", "Fecha_sort", "Tipo"]
    ).drop(columns=["orden_alerta", "orden_prioridad", "Fecha_sort"])
    return df_recordatorios


def construir_pendientes_prioritarios(
    df_nc,
    df_acciones,
    df_documentos,
    df_auditorias=None,
    df_hallazgos=None,
    df_versiones=None,
):
    pendientes = []

    if not df_nc.empty:
        pendientes_nc = df_nc[
            (df_nc["Estado"] != "Cerrada")
            & (df_nc["Severidad"].isin(["Alta", "Crítica"]))
        ]
        for _, fila in pendientes_nc.iterrows():
            pendientes.append(
                {
                    "Categoría": "NC crítica",
                    "Referencia": fila["Código"],
                    "Detalle": fila["Título"],
                    "Responsable": fila["Responsable"],
                    "Estado": fila["Estado"],
                    "Fecha crítica": fila["Fecha compromiso"],
                    "Prioridad": "Crítica" if fila["Severidad"] == "Crítica" else "Alta",
                    "Motivo": "No conformidad abierta de severidad alta o crítica.",
                }
            )

        pendientes_aprobacion_nc = df_nc[
            (df_nc["Fecha cierre"].notna())
            & (df_nc["Aprobado por"].isna() | (df_nc["Aprobado por"].astype(str).str.strip() == ""))
        ]
        for _, fila in pendientes_aprobacion_nc.iterrows():
            pendientes.append(
                {
                    "Categoría": "NC pendiente de aprobación",
                    "Referencia": fila["Código"],
                    "Detalle": fila["Título"],
                    "Responsable": fila["Responsable"],
                    "Estado": fila["Estado"],
                    "Fecha crítica": fila["Fecha cierre"],
                    "Prioridad": "Alta",
                    "Motivo": "La no conformidad tiene cierre registrado pero falta aprobación formal.",
                }
            )

    if not df_acciones.empty:
        acciones_vencidas = df_acciones[
            (df_acciones["Estado"] != "Cerrada")
            & (pd.to_datetime(df_acciones["Fecha compromiso"], errors="coerce") < pd.Timestamp.today().normalize())
        ]
        for _, fila in acciones_vencidas.iterrows():
            pendientes.append(
                {
                    "Categoría": "Acción vencida",
                    "Referencia": fila["Título"],
                    "Detalle": fila["Código NC"],
                    "Responsable": fila["Responsable"],
                    "Estado": fila["Estado"],
                    "Fecha crítica": fila["Fecha compromiso"],
                    "Prioridad": "Crítica",
                    "Motivo": "La acción correctiva o preventiva ya está fuera de plazo.",
                }
            )

        pendientes_aprobacion_acc = df_acciones[
            (df_acciones["Fecha cierre"].notna())
            & (df_acciones["Aprobado por"].isna() | (df_acciones["Aprobado por"].astype(str).str.strip() == ""))
        ]
        for _, fila in pendientes_aprobacion_acc.iterrows():
            pendientes.append(
                {
                    "Categoría": "Acción pendiente de cierre",
                    "Referencia": fila["Título"],
                    "Detalle": fila["Código NC"],
                    "Responsable": fila["Responsable"],
                    "Estado": fila["Estado"],
                    "Fecha crítica": fila["Fecha cierre"],
                    "Prioridad": "Alta",
                    "Motivo": "La acción ya reportó cierre pero todavía no cuenta con aprobación formal.",
                }
            )

    if not df_documentos.empty:
        docs_borrador = df_documentos[df_documentos["Estado"] == "Borrador"]
        for _, fila in docs_borrador.iterrows():
            pendientes.append(
                {
                    "Categoría": "Documento borrador",
                    "Referencia": fila["Código"],
                    "Detalle": fila["Nombre"],
                    "Responsable": fila["Aprobado por"],
                    "Estado": fila["Estado"],
                    "Fecha crítica": fila["Vigente desde"],
                    "Prioridad": "Media",
                    "Motivo": "Documento pendiente de formalizar y liberar como vigente.",
                }
            )

        docs_vencidos = df_documentos[
            (df_documentos["Estado"] == "Vigente")
            & (pd.to_datetime(df_documentos["Vigente hasta"], errors="coerce") < pd.Timestamp.today().normalize())
        ]
        for _, fila in docs_vencidos.iterrows():
            pendientes.append(
                {
                    "Categoría": "Documento vencido",
                    "Referencia": fila["Código"],
                    "Detalle": fila["Nombre"],
                    "Responsable": fila["Aprobado por"],
                    "Estado": fila["Estado"],
                    "Fecha crítica": fila["Vigente hasta"],
                    "Prioridad": "Alta",
                    "Motivo": "Documento vigente fuera de fecha de vigencia.",
                }
            )

    if df_auditorias is not None and not df_auditorias.empty:
        auditorias_proximas = df_auditorias[
            (df_auditorias["Estado"] != "Cerrada")
            & (
                pd.to_datetime(df_auditorias["Fecha programada"], errors="coerce")
                <= (pd.Timestamp.today().normalize() + pd.Timedelta(days=14))
            )
        ]
        for _, fila in auditorias_proximas.iterrows():
            pendientes.append(
                {
                    "Categoría": "Auditoría prioritaria",
                    "Referencia": fila["Código"],
                    "Detalle": fila["Título"],
                    "Responsable": fila["Auditor líder"],
                    "Estado": fila["Estado"],
                    "Fecha crítica": fila["Fecha programada"],
                    "Prioridad": "Alta" if clasificar_vencimiento(fila["Fecha programada"], 14) in ["Vencido", "Vence hoy"] else "Media",
                    "Motivo": "Auditoría próxima que requiere preparación o cierre.",
                }
            )

    if df_hallazgos is not None and not df_hallazgos.empty:
        hallazgos_prioritarios = df_hallazgos[
            (df_hallazgos["Estado"] != "Cerrado")
            & (df_hallazgos["Severidad"].isin(["Alta", "Crítica"]))
        ]
        for _, fila in hallazgos_prioritarios.iterrows():
            pendientes.append(
                {
                    "Categoría": "Hallazgo prioritario",
                    "Referencia": fila["Referencia"],
                    "Detalle": fila["Descripción"],
                    "Responsable": fila["Responsable"],
                    "Estado": fila["Estado"],
                    "Fecha crítica": fila["Fecha compromiso"],
                    "Prioridad": "Crítica" if fila["Severidad"] == "Crítica" else "Alta",
                    "Motivo": "Hallazgo de auditoría de severidad alta o crítica aún abierto.",
                }
            )

    if df_versiones is not None and not df_versiones.empty:
        versiones_pendientes = df_versiones[
            df_versiones["Aprobado por"].isna() | (df_versiones["Aprobado por"].astype(str).str.strip() == "")
        ]
        for _, fila in versiones_pendientes.iterrows():
            pendientes.append(
                {
                    "Categoría": "Versión pendiente",
                    "Referencia": fila["Código"],
                    "Detalle": f"{fila['Nombre']} v{fila['Versión']}",
                    "Responsable": fila["Elaborado por"],
                    "Estado": "Pendiente",
                    "Fecha crítica": fila["Registrada el"],
                    "Prioridad": "Alta",
                    "Motivo": "Versión documental pendiente de aprobación.",
                }
            )

    if not pendientes:
        return pd.DataFrame(
            columns=["Categoría", "Referencia", "Detalle", "Responsable", "Estado", "Fecha crítica", "Prioridad", "Motivo"]
        )

    df_pendientes = pd.DataFrame(pendientes)
    orden_prioridad = {"Crítica": 0, "Alta": 1, "Media": 2, "Baja": 3}
    df_pendientes["orden_prioridad"] = df_pendientes["Prioridad"].map(orden_prioridad).fillna(99)
    df_pendientes["fecha_sort"] = pd.to_datetime(df_pendientes["Fecha crítica"], errors="coerce")
    df_pendientes = df_pendientes.sort_values(by=["orden_prioridad", "fecha_sort", "Categoría"]).drop(
        columns=["orden_prioridad", "fecha_sort"]
    )
    return df_pendientes


def calcular_promedio_cierre(df, fecha_inicio, fecha_fin):
    if df.empty or fecha_inicio not in df.columns or fecha_fin not in df.columns:
        return 0.0

    tiempos = df[[fecha_inicio, fecha_fin]].copy()
    tiempos[fecha_inicio] = pd.to_datetime(tiempos[fecha_inicio], errors="coerce")
    tiempos[fecha_fin] = pd.to_datetime(tiempos[fecha_fin], errors="coerce")
    tiempos = tiempos.dropna()

    if tiempos.empty:
        return 0.0

    tiempos = tiempos[tiempos[fecha_fin] >= tiempos[fecha_inicio]]
    if tiempos.empty:
        return 0.0

    return round((tiempos[fecha_fin] - tiempos[fecha_inicio]).dt.days.mean(), 1)


def construir_serie_mensual(df, fecha_columna, etiqueta, meses=6):
    if fecha_columna not in df.columns:
        return pd.DataFrame(columns=["Mes", etiqueta])

    fechas = pd.to_datetime(df[fecha_columna], errors="coerce").dropna()
    hoy = pd.Timestamp.today().normalize()
    inicio = (hoy - pd.DateOffset(months=meses - 1)).to_period("M")
    fin = hoy.to_period("M")
    indice = pd.period_range(start=inicio, end=fin, freq="M")

    if fechas.empty:
        return pd.DataFrame({"Mes": indice.astype(str), etiqueta: [0] * len(indice)})

    serie = fechas.dt.to_period("M").value_counts().sort_index()
    serie = serie.reindex(indice, fill_value=0)
    return pd.DataFrame({"Mes": indice.astype(str), etiqueta: serie.values})


def combinar_series_mensuales(df_izq, df_der, columna_mes="Mes"):
    if df_izq.empty:
        return df_der
    if df_der.empty:
        return df_izq
    return df_izq.merge(df_der, on=columna_mes, how="outer").fillna(0)


def formatear_porcentaje(parte, total):
    if not total:
        return "0%"
    return f"{round((parte / total) * 100, 1)}%"


def generar_excel_ejecutivo(secciones):
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for nombre, dataframe in secciones.items():
            if dataframe is None:
                continue
            df = dataframe.copy()
            hoja = nombre[:31]
            df.to_excel(writer, sheet_name=hoja, index=False)
            worksheet = writer.sheets[hoja]
            for col_idx, column_cells in enumerate(worksheet.columns, start=1):
                max_length = 0
                for cell in column_cells:
                    cell_value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(cell_value))
                worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 14), 40)
    return buffer.getvalue()


def generar_pdf_ejecutivo(resumen_lineas, secciones):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise RuntimeError("Instala reportlab para exportar a PDF.") from exc

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elementos = [Paragraph("Reporte ejecutivo de calidad", styles["Title"]), Spacer(1, 12)]

    for linea in resumen_lineas:
        elementos.append(Paragraph(linea, styles["BodyText"]))
        elementos.append(Spacer(1, 6))

    for nombre, dataframe in secciones.items():
        elementos.append(Spacer(1, 10))
        elementos.append(Paragraph(nombre, styles["Heading2"]))
        df = dataframe.copy()
        if df.empty:
            elementos.append(Paragraph("Sin datos en esta sección.", styles["BodyText"]))
            continue

        df = df.head(15)
        tabla_data = [df.columns.tolist()] + df.fillna("").astype(str).values.tolist()
        tabla = Table(tabla_data, repeatRows=1)
        tabla.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d9edf2")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#16324f")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#c7d4de")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fbfc")]),
                ]
            )
        )
        elementos.append(tabla)

    doc.build(elementos)
    return buffer.getvalue()

def pantalla_acceso():
    st.title("Acceso al sistema")
    st.caption(f"Persistencia de usuarios: {obtener_backend_usuarios()}")
    pestana_login, pestana_registro = st.tabs(["Iniciar sesión", "Crear cuenta"])

    with pestana_login:
        email_login = st.text_input("Correo", key="login_email")
        password_login = st.text_input("Contraseña", type="password", key="login_password")

        if st.button("Ingresar"):
            try:
                resultado = autenticar_usuario(email_login, password_login)

                if resultado["ok"]:
                    st.session_state["autenticado"] = True
                    st.session_state["usuario_email"] = resultado["email"]
                    st.session_state["es_admin"] = resultado["es_admin"]
                    st.session_state["rol_usuario"] = resultado.get("rol", "captura")
                    st.rerun()
                else:
                    st.error(resultado["mensaje"])
            except Exception as e:
                st.error(f"Error al iniciar sesión: {e}")

    with pestana_registro:
        email_registro = st.text_input("Correo institucional o personal", key="registro_email")
        password_registro = st.text_input("Contraseña", type="password", key="registro_password")
        password_registro_2 = st.text_input("Confirmar contraseña", type="password", key="registro_password_2")
        rol_registro = st.selectbox(
            "Perfil solicitado",
            ["captura", "responsable", "auditor", "calidad"],
            format_func=lambda valor: valor.capitalize(),
            key="registro_rol",
        )

        if st.button("Crear cuenta"):
            try:
                if not email_registro or not password_registro:
                    st.warning("Completa correo y contraseña.")
                elif password_registro != password_registro_2:
                    st.warning("Las contraseñas no coinciden.")
                else:
                    registrar_usuario(email_registro, password_registro, rol_registro)
                    st.success("Cuenta creada. Queda pendiente de aprobación.")
            except Exception as e:
                st.error(f"No se pudo crear la cuenta: {e}")


if "autenticado" not in st.session_state:
    st.session_state["autenticado"] = False

if "es_admin" not in st.session_state:
    st.session_state["es_admin"] = False

if "rol_usuario" not in st.session_state:
    st.session_state["rol_usuario"] = "captura"

if not st.session_state["autenticado"]:
    pantalla_acceso()
    st.stop()


def mostrar_biobanco():
    st.title("Sistema de Seguimiento de Voluntarios y Biobanco")

    st.subheader("Dashboard resumen")

    col1, col2, col3 = st.columns(3)

    with col1:
        total_voluntarios = contar_voluntarios()
        st.metric("Voluntarios registrados", total_voluntarios)

    with col2:
        total_pendientes = contar_visitas_pendientes()
        st.metric("Visitas pendientes", total_pendientes)

    with col3:
        total_racks = contar_racks_activos()
        st.metric("Racks activos", total_racks)

    try:
        ocupacion = obtener_ocupacion_racks()

        if ocupacion:
            st.subheader("Ocupación de racks")
            df_ocupacion = pd.DataFrame(
                ocupacion,
                columns=["ID Rack", "Tipo", "Capacidad", "Ocupadas"]
            )
            df_ocupacion["% Ocupación"] = (
                df_ocupacion["Ocupadas"] / df_ocupacion["Capacidad"] * 100
            ).round(2)

            st.dataframe(df_ocupacion, use_container_width=True)

    except Exception as e:
        st.warning(f"No se pudo mostrar el dashboard: {e}")

    st.subheader("Registro de voluntarios")

    with st.form("form_voluntario"):
        col1, col2 = st.columns(2)

        with col1:
            id_voluntario = st.text_input("ID del voluntario")
            expediente = st.text_input("Expediente")
            fecha_toma1 = st.date_input("Fecha Toma 1")
            apellido_paterno = st.text_input("Apellido paterno")
            apellido_materno = st.text_input("Apellido materno")
            nombre = st.text_input("Nombre")
            genero = st.selectbox("Género", ["M", "F"])
            residencia = st.text_input("Residencia")
            fecha_nacimiento = st.date_input("Fecha de nacimiento")

        with col2:
            edad = st.number_input("Edad", min_value=0, max_value=120, step=1)
            peso = st.number_input("Peso", min_value=0.0, step=0.1)
            estatura = st.number_input("Estatura", min_value=0.0, step=0.01)
            tubos_amarillos = st.number_input("Tubos amarillos", min_value=0, step=1)
            tubos_verdes = st.number_input("Tubos verdes", min_value=0, step=1)
            cantidad_pbmc = st.number_input("Número de alícuotas PBMC", min_value=0, step=1)
            conteo_celular = st.text_input("Conteo celular PBMC")
            correo = st.text_input("Correo")
            telefono = st.text_input("Teléfono")
            patologias = st.text_area("Patologías")
            observaciones = st.text_area("Observaciones")

        enviado = st.form_submit_button("Registrar voluntario")

    if enviado:
        datos = {
            "id_voluntario": id_voluntario,
            "expediente": expediente,
            "fecha_toma1": str(fecha_toma1),
            "apellido_paterno": apellido_paterno,
            "apellido_materno": apellido_materno,
            "nombre": nombre,
            "genero": genero,
            "residencia": residencia,
            "fecha_nacimiento": str(fecha_nacimiento),
            "edad": edad,
            "peso": peso,
            "estatura": estatura,
            "tubos_amarillos": tubos_amarillos,
            "tubos_verdes": tubos_verdes,
            "cantidad_pbmc": cantidad_pbmc,
            "conteo_celular": conteo_celular,
            "correo": correo,
            "telefono": telefono,
            "patologias": patologias,
            "observaciones": observaciones,
        }

        try:
            registrar_voluntario(datos)
            st.success(f"Voluntario {id_voluntario} registrado correctamente.")

            visitas = ver_visitas(id_voluntario)
            st.subheader("Visitas generadas")
            st.table(visitas)

        except Exception as e:
            st.error(f"Error al registrar: {e}")

    st.subheader("Racks registrados")

    try:
        racks = ver_racks()
        if racks:
            df_racks = pd.DataFrame(
                racks,
                columns=["ID Rack", "Tipo", "Capacidad", "Ocupadas"]
            )
            st.dataframe(df_racks, use_container_width=True)
        else:
            st.info("No hay racks registrados todavía.")
    except Exception as e:
        st.warning(f"No se pudieron mostrar los racks: {e}")

    st.subheader("Visualización de racks 9x9")
    col1, col2 = st.columns(2)

    with col1:
        st.write("### Suero")
        try:
            rack_suero = ver_rack_suero("SUERO_1")
            st.table(rack_suero)
        except Exception:
            st.info("No hay datos de suero todavía")

    with col2:
        st.write("### PBMC")
        try:
            rack_pbmc = ver_rack_pbmc("PBMC_1")
            st.table(rack_pbmc)
        except Exception:
            st.info("No hay datos de PBMC todavía")

    st.subheader("Exportar información")

    if st.button("Generar archivo Excel"):
        try:
            archivo = exportar_a_excel()

            with open(archivo, "rb") as f:
                st.download_button(
                    label="Descargar reporte Excel",
                    data=f,
                    file_name=archivo,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as e:
            st.error(f"Error al exportar: {e}")

    st.subheader("Consulta por ID de voluntario")

    id_busqueda = st.text_input("Ingresa el ID del voluntario a consultar")

    if st.button("Buscar voluntario"):
        try:
            voluntario = buscar_voluntario_por_id(id_busqueda)

            if voluntario:
                columnas_voluntario = [
                    "id_voluntario", "expediente", "fecha_toma1",
                    "apellido_paterno", "apellido_materno", "nombre",
                    "genero", "residencia", "fecha_nacimiento", "edad",
                    "peso", "estatura", "tubos_amarillos", "tubos_verdes",
                    "correo", "telefono", "patologias", "observaciones",
                ]

                df_voluntario = pd.DataFrame([voluntario], columns=columnas_voluntario)
                st.write("### Datos del voluntario")
                st.dataframe(df_voluntario, use_container_width=True)

                visitas = ver_visitas(id_busqueda)
                if visitas:
                    df_visitas = pd.DataFrame(
                        visitas,
                        columns=["Tipo toma", "Fecha programada", "Fecha real", "Estado"],
                    )
                    st.write("### Visitas")
                    st.dataframe(df_visitas, use_container_width=True)

                suero = ver_alicuotas_suero_voluntario(id_busqueda)
                if suero:
                    df_suero = pd.DataFrame(
                        suero,
                        columns=["Núm alícuota", "Fecha ingreso", "Rack", "Fila", "Columna", "Tipo toma"],
                    )
                    st.write("### Alícuotas de suero")
                    st.dataframe(df_suero, use_container_width=True)

                pbmc = ver_alicuotas_pbmc_voluntario(id_busqueda)
                if pbmc:
                    df_pbmc = pd.DataFrame(
                        pbmc,
                        columns=["Núm alícuota", "Conteo celular", "Fecha ingreso", "Rack", "Fila", "Columna", "Tipo toma"],
                    )
                    st.write("### Alícuotas PBMC")
                    st.dataframe(df_pbmc, use_container_width=True)

            else:
                st.warning("No se encontró un voluntario con ese ID.")

        except Exception as e:
            st.error(f"Error en la búsqueda: {e}")


def mostrar_proyecto_hospitalario():
    st.markdown(
        """
        <div class="iner-hero">
            <div class="iner-kicker">Proyecto Hospitalario</div>
            <h1 class="iner-title">Seguimiento hospitalario y tomas pendientes</h1>
            <p class="iner-copy">Vista operativa del Google Sheet con búsqueda clínica, edición por toma y resúmenes listos para revisión e impresión.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    try:
        df_sheet = leer_sheet_como_dataframe()
        df_sheet = preparar_datos_hospitalarios(df_sheet)
        df_sheet = preparar_resumen_influenza_observaciones(df_sheet)

        if "CLAVE DE LABORATORIO" not in df_sheet.columns:
            st.error("La hoja no contiene la columna 'CLAVE DE LABORATORIO'.")
            return

        tarjeta_seccion(
            "Consulta",
            "Búsqueda por clave de laboratorio",
            "Encuentra rápidamente registros clínicos por clave para inspección o seguimiento.",
        )
        clave_busqueda = st.text_input("Ingresa la clave de laboratorio", placeholder="Ej. VSR_H_001")

        if clave_busqueda:
            df_filtrado_clave = df_sheet[
                df_sheet["CLAVE DE LABORATORIO"].astype(str).str.contains(clave_busqueda, case=False, na=False)
            ]

            if not df_filtrado_clave.empty:
                st.write("### Resultado de búsqueda")
                st.dataframe(df_filtrado_clave, use_container_width=True)
            else:
                st.warning("No se encontró ningún paciente con esa clave de laboratorio.")

        tarjeta_seccion(
            "Edición",
            "Edición dinámica de tomas",
            "Carga un paciente por clave y modifica solo las columnas asociadas a la toma seleccionada.",
        )

        clave_edicion = st.text_input(
            "Clave de laboratorio para editar",
            placeholder="Ej. VSR_H_001",
            key="clave_edicion",
        )

        toma_seleccionada = st.selectbox(
            "Selecciona la toma",
            [f"T{i}" for i in range(1, 15)],
        )

        if st.button("Cargar paciente"):
            try:
                numero_fila = buscar_fila_por_clave(clave_edicion)

                if numero_fila:
                    fila_paciente = obtener_fila_como_diccionario(numero_fila)

                    st.session_state["fila_paciente"] = fila_paciente
                    st.session_state["clave_edicion_actual"] = clave_edicion

                    st.success(f"Paciente encontrado en fila {numero_fila}")
                else:
                    st.warning("No se encontró la clave")

            except Exception as e:
                st.error(f"Error al cargar paciente: {e}")

        if "fila_paciente" in st.session_state:
            fila_paciente = st.session_state["fila_paciente"]
            clave_actual = st.session_state["clave_edicion_actual"]

            numero_toma = int(toma_seleccionada.replace("T", ""))
            columnas_toma = obtener_columnas_por_toma(numero_toma)

            st.write(f"### Editando {toma_seleccionada}")

            if columnas_toma:
                cambios = {}

                col1, col2 = st.columns(2)

                for i, col in enumerate(columnas_toma):
                    valor_actual = fila_paciente.get(col, "")

                    with col1 if i % 2 == 0 else col2:
                        nuevo_valor = st.text_input(
                            col,
                            value=valor_actual,
                            key=f"{col}_{numero_toma}",
                        )

                    cambios[col] = nuevo_valor

                if st.button("Guardar cambios"):
                    try:
                        actualizar_campos_por_clave(clave_actual, cambios)
                        st.success("Cambios guardados correctamente en Google Sheets")

                        numero_fila = buscar_fila_por_clave(clave_actual)
                        fila_actualizada = obtener_fila_como_diccionario(numero_fila)
                        st.session_state["fila_paciente"] = fila_actualizada

                    except Exception as e:
                        st.error(f"Error al guardar: {e}")
            else:
                st.warning(f"No se encontraron columnas para {toma_seleccionada}")

        tarjeta_seccion("Resumen", "Distribución por diagnóstico")

        if "DIAGNOSTICO_GRUPO" in df_sheet.columns:
            conteo_diag = df_sheet["DIAGNOSTICO_GRUPO"].value_counts().reset_index()
            conteo_diag.columns = ["Diagnóstico agrupado", "Número de pacientes"]
            st.dataframe(conteo_diag, use_container_width=True)

        tarjeta_seccion("Resumen", "Distribución por grupo de edad")

        if "GRUPO_EDAD" in df_sheet.columns:
            conteo_edad = df_sheet["GRUPO_EDAD"].value_counts().reset_index()
            conteo_edad.columns = ["Grupo de edad", "Número de pacientes"]
            st.dataframe(conteo_edad, use_container_width=True)

        tarjeta_seccion("Cruce", "Subclasificación de edad y diagnóstico")

        if "GRUPO_EDAD" in df_sheet.columns and "DIAGNOSTICO_GRUPO" in df_sheet.columns:
            tabla_cruzada = pd.crosstab(
                df_sheet["GRUPO_EDAD"],
                df_sheet["DIAGNOSTICO_GRUPO"],
            )

            orden_edades = ["Bebé", "Niño", "Adolescente", "Adulto"]
            tabla_cruzada = tabla_cruzada.reindex(
                [edad for edad in orden_edades if edad in tabla_cruzada.index]
            )

            columnas_preferidas = [
                "COVID",
                "Coinfección COVID",
                "INFLUENZA",
                "Coinfección INFLUENZA",
                "VSR",
                "Coinfección VSR",
            ]

            columnas_existentes = list(tabla_cruzada.columns)
            columnas_extra = [col for col in columnas_existentes if col not in columnas_preferidas]
            columnas_finales = [col for col in columnas_preferidas if col in columnas_existentes] + columnas_extra

            tabla_cruzada = tabla_cruzada[columnas_finales]

            st.dataframe(tabla_cruzada, use_container_width=True)

        tarjeta_seccion(
            "Vigilancia",
            "Resumen de influenza y coinfecciones en observaciones",
            "Detección automática basada en el contenido de observaciones del Google Sheet.",
        )

        if "INFLUENZA_OBS_GRUPO" in df_sheet.columns:
            df_influenza = df_sheet[df_sheet["INFLUENZA_OBS_GRUPO"].notna()]

            if not df_influenza.empty:
                conteo_influenza = df_influenza["INFLUENZA_OBS_GRUPO"].value_counts().reset_index()
                conteo_influenza.columns = ["Influenza / coinfección detectada", "Número de pacientes"]
                st.dataframe(conteo_influenza, use_container_width=True)
            else:
                st.info("No se encontraron registros de influenza en observaciones.")

        df_tomas_pendientes = construir_tabla_tomas_pendientes(df_sheet)
        tarjeta_seccion(
            "Operación",
            "Tomas pendientes del día",
            "Lista operativa con folio, nombre, observaciones, toma pendiente y clave de laboratorio, lista para exportación.",
        )

        if not df_tomas_pendientes.empty:
            st.dataframe(df_tomas_pendientes, use_container_width=True)

            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                df_tomas_pendientes.to_excel(writer, sheet_name="Tomas pendientes", index=False)
                worksheet = writer.sheets["Tomas pendientes"]

                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1):
                    for cell in row:
                        cell.number_format = "@"
                        cell.value = str(cell.value) if cell.value is not None else ""

                for col_idx, column_cells in enumerate(worksheet.columns, start=1):
                    max_length = 0

                    for cell in column_cells:
                        cell_value = "" if cell.value is None else str(cell.value)
                        max_length = max(max_length, len(cell_value))

                    adjusted_width = min(max(max_length + 2, 14), 40)
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

            st.download_button(
                label="Descargar tomas pendientes en Excel",
                data=buffer.getvalue(),
                file_name="tomas_pendientes_lit.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.info("No se encontraron tomas pendientes en columnas de ingreso a LIT.")

        tarjeta_seccion("Base", "Datos completos del Google Sheet")
        st.dataframe(df_sheet, use_container_width=True)

        tarjeta_seccion("Base", "Resumen general de pacientes")
        df_resumen_pacientes = construir_tabla_resumen_pacientes(df_sheet)
        st.dataframe(df_resumen_pacientes, use_container_width=True)

    except Exception as e:
        st.error(f"Error al leer Google Sheets: {e}")


def mostrar_calidad():
    def filtrar_dataframe(df, filtros):
        df_filtrado = df.copy()
        for columna, valor in filtros.items():
            if valor and valor != "Todos" and columna in df_filtrado.columns:
                df_filtrado = df_filtrado[df_filtrado[columna].astype(str) == valor]
        return df_filtrado

    st.markdown(
        """
        <div class="iner-hero">
            <div class="iner-kicker">Sistema de calidad</div>
            <h1 class="iner-title">Seguimiento de no conformidades y acciones</h1>
            <p class="iner-copy">Espacio operativo para apoyar la trazabilidad requerida en validación ISO 9001 dentro de la misma plataforma.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.caption(
        f"Perfil actual: {rol_actual().capitalize()} | "
        "Captura registra información, Responsable da seguimiento, Auditor gestiona auditorías, "
        "Calidad aprueba cierres y documentos, Admin conserva control total."
    )
    st.caption(f"Persistencia de Calidad: {obtener_backend_calidad()}")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("No conformidades abiertas", contar_no_conformidades_abiertas())
    with col2:
        st.metric("Acciones abiertas", contar_acciones_abiertas())
    with col3:
        st.metric("Acciones vencidas", contar_acciones_vencidas())

    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs(["No conformidades", "Acciones", "Auditorías", "Control documental", "Seguimiento", "Indicadores", "Respaldo"])

    with tab1:
        tarjeta_seccion(
            "Registro",
            "Nueva no conformidad",
            "Documenta hallazgos, desviaciones, incidentes o incumplimientos con responsable y fecha compromiso.",
        )

        enviado_nc = False
        if tiene_rol("captura", "responsable", "calidad"):
            with st.form("form_no_conformidad", clear_on_submit=True):
                col1, col2 = st.columns(2)

                with col1:
                    codigo = st.text_input("Código", placeholder="NC-2026-001")
                    titulo = st.text_input("Título")
                    origen = st.selectbox("Origen", ["Auditoría", "Proceso", "Cliente", "Proveedor", "Interno"])
                    area = st.text_input("Área o proceso")
                    severidad = st.selectbox("Severidad", ["Baja", "Media", "Alta", "Crítica"])
                    fecha_deteccion = st.date_input("Fecha de detección")

                with col2:
                    detectado_por = st.text_input("Detectado por")
                    responsable = st.text_input("Responsable")
                    fecha_compromiso = st.date_input("Fecha compromiso")
                    descripcion = st.text_area("Descripción")
                    causa_raiz = st.text_area("Causa raíz")

                enviado_nc = st.form_submit_button("Guardar no conformidad")

            if enviado_nc:
                datos_nc = {
                    "codigo": codigo.strip(),
                    "titulo": titulo.strip(),
                    "descripcion": descripcion.strip(),
                    "origen": origen,
                    "area": area.strip(),
                    "severidad": severidad,
                    "detectado_por": detectado_por.strip(),
                    "responsable": responsable.strip(),
                    "fecha_deteccion": str(fecha_deteccion),
                    "fecha_compromiso": str(fecha_compromiso),
                    "causa_raiz": causa_raiz.strip(),
                    "usuario_email": st.session_state.get("usuario_email", ""),
                }

                campos_requeridos = [
                    "codigo",
                    "titulo",
                    "descripcion",
                    "area",
                    "detectado_por",
                    "responsable",
                ]

                if any(not datos_nc[campo] for campo in campos_requeridos):
                    st.warning("Completa todos los campos obligatorios de la no conformidad.")
                else:
                    try:
                        registrar_no_conformidad(datos_nc)
                        st.success("No conformidad registrada correctamente.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"No se pudo registrar la no conformidad: {e}")
        else:
            mostrar_aviso_permiso("solo Captura, Responsable, Calidad o Admin pueden registrar no conformidades.")

        no_conformidades = listar_no_conformidades()
        st.subheader("No conformidades registradas")

        if no_conformidades:
            df_nc = pd.DataFrame(
                no_conformidades,
                columns=[
                    "ID", "Código", "Título", "Descripción", "Origen", "Área", "Severidad",
                    "Estado", "Detectado por", "Responsable", "Fecha detección",
                    "Fecha compromiso", "Causa raíz", "Fecha cierre", "Aprobado por", "Fecha aprobación", "Comentario final",
                ],
            )

            colf1, colf2, colf3 = st.columns(3)
            filtro_estado_nc = colf1.selectbox("Filtrar por estado", ["Todos"] + sorted(df_nc["Estado"].dropna().unique().tolist()), key="filtro_estado_nc")
            filtro_area_nc = colf2.selectbox("Filtrar por área", ["Todos"] + sorted(df_nc["Área"].dropna().unique().tolist()), key="filtro_area_nc")
            filtro_responsable_nc = colf3.selectbox("Filtrar por responsable", ["Todos"] + sorted(df_nc["Responsable"].dropna().unique().tolist()), key="filtro_responsable_nc")

            df_nc_filtrado = filtrar_dataframe(
                df_nc,
                {
                    "Estado": filtro_estado_nc,
                    "Área": filtro_area_nc,
                    "Responsable": filtro_responsable_nc,
                },
            )
            st.dataframe(df_nc_filtrado, use_container_width=True, hide_index=True)
        else:
            st.info("Todavía no hay no conformidades registradas.")

        tarjeta_seccion(
            "Edición",
            "Editar no conformidad",
            "Permite corregir o completar datos base sin perder la trazabilidad, ya que cada edición entra en bitácora.",
        )

        if no_conformidades and tiene_rol("responsable", "calidad"):
            nc_por_label = {
                f"{codigo} | {titulo}": {
                    "id": id_nc,
                    "codigo": codigo,
                    "titulo": titulo,
                    "descripcion": descripcion,
                    "origen": origen,
                    "area": area,
                    "severidad": severidad,
                    "estado": estado,
                    "detectado_por": detectado_por,
                    "responsable": responsable,
                    "fecha_deteccion": fecha_deteccion,
                    "fecha_compromiso": fecha_compromiso,
                    "causa_raiz": causa_raiz,
                    "fecha_cierre": fecha_cierre,
                    "aprobado_por": aprobado_por,
                    "fecha_aprobacion": fecha_aprobacion,
                    "comentario_final": comentario_final,
                }
                for id_nc, codigo, titulo, descripcion, origen, area, severidad, estado, detectado_por, responsable, fecha_deteccion, fecha_compromiso, causa_raiz, fecha_cierre, aprobado_por, fecha_aprobacion, comentario_final
                in no_conformidades
            }

            with st.form("form_editar_nc"):
                seleccion_editar_nc = st.selectbox("No conformidad a editar", list(nc_por_label.keys()))
                nc_actual = nc_por_label[seleccion_editar_nc]
                col1, col2 = st.columns(2)
                with col1:
                    editar_titulo_nc = st.text_input("Título", value=nc_actual["titulo"], key="editar_titulo_nc")
                    editar_origen_nc = st.selectbox("Origen", ["Auditoría", "Proceso", "Cliente", "Proveedor", "Interno"], index=["Auditoría", "Proceso", "Cliente", "Proveedor", "Interno"].index(nc_actual["origen"]), key="editar_origen_nc")
                    editar_area_nc = st.text_input("Área", value=nc_actual["area"], key="editar_area_nc")
                    editar_severidad_nc = st.selectbox("Severidad", ["Baja", "Media", "Alta", "Crítica"], index=["Baja", "Media", "Alta", "Crítica"].index(nc_actual["severidad"]), key="editar_severidad_nc")
                    editar_detectado_por_nc = st.text_input("Detectado por", value=nc_actual["detectado_por"], key="editar_detectado_por_nc")
                with col2:
                    editar_responsable_nc = st.text_input("Responsable", value=nc_actual["responsable"], key="editar_responsable_nc")
                    editar_fecha_deteccion_nc = st.text_input("Fecha detección", value=nc_actual["fecha_deteccion"], key="editar_fecha_deteccion_nc")
                    editar_fecha_compromiso_nc = st.text_input("Fecha compromiso", value=nc_actual["fecha_compromiso"] or "", key="editar_fecha_compromiso_nc")
                    editar_causa_raiz_nc = st.text_area("Causa raíz", value=nc_actual["causa_raiz"] or "", key="editar_causa_raiz_nc")
                    editar_descripcion_nc = st.text_area("Descripción", value=nc_actual["descripcion"] or "", key="editar_descripcion_nc")
                guardar_edicion_nc = st.form_submit_button("Guardar edición")

            if guardar_edicion_nc:
                try:
                    actualizar_no_conformidad(
                        {
                            "id_no_conformidad": nc_actual["id"],
                            "codigo": nc_actual["codigo"],
                            "titulo": editar_titulo_nc.strip(),
                            "descripcion": editar_descripcion_nc.strip(),
                            "origen": editar_origen_nc,
                            "area": editar_area_nc.strip(),
                            "severidad": editar_severidad_nc,
                            "detectado_por": editar_detectado_por_nc.strip(),
                            "responsable": editar_responsable_nc.strip(),
                            "fecha_deteccion": editar_fecha_deteccion_nc.strip(),
                            "fecha_compromiso": editar_fecha_compromiso_nc.strip(),
                            "causa_raiz": editar_causa_raiz_nc.strip(),
                            "usuario_email": st.session_state.get("usuario_email", ""),
                        }
                    )
                    st.success("No conformidad editada correctamente.")
                    st.rerun()
                except Exception as e:
                    st.error(f"No se pudo editar la no conformidad: {e}")

        elif no_conformidades:
            mostrar_aviso_permiso("solo Responsable, Calidad o Admin pueden editar no conformidades.")

        tarjeta_seccion(
            "Estado",
            "Actualizar no conformidad",
            "Puedes mover hallazgos a En proceso y solo Calidad o Admin pueden cerrarlos formalmente.",
        )

        if no_conformidades:
            opciones_estado_nc = {
                f"{codigo} | {titulo} | {estado}": id_nc
                for id_nc, codigo, titulo, _, _, _, _, estado, *_ in no_conformidades
            }

            if tiene_rol("responsable", "calidad"):
                with st.form("form_estado_nc"):
                    seleccion_nc = st.selectbox("Selecciona la no conformidad", list(opciones_estado_nc.keys()))
                    nuevo_estado_nc = st.selectbox("Nuevo estado", ["Abierta", "En proceso"])
                    causa_raiz_update = st.text_area("Actualizar causa raíz", key="causa_raiz_update")
                    verificacion_cierre = st.text_area("Verificación de cierre", key="verificacion_cierre_nc")
                    guardar_estado_nc = st.form_submit_button("Actualizar estado")

                if guardar_estado_nc:
                    try:
                        actualizar_estado_no_conformidad(
                            id_no_conformidad=opciones_estado_nc[seleccion_nc],
                            nuevo_estado=nuevo_estado_nc,
                            causa_raiz=causa_raiz_update.strip() or None,
                            verificacion_cierre=verificacion_cierre.strip() or None,
                            es_admin=st.session_state.get("es_admin", False),
                            rol_usuario=rol_actual(),
                            usuario_email=st.session_state.get("usuario_email", ""),
                        )
                        st.success("Estado de no conformidad actualizado.")
                        st.rerun()
                    except PermissionError as e:
                        st.warning(str(e))
                    except Exception as e:
                        st.error(f"No se pudo actualizar el estado: {e}")
            else:
                mostrar_aviso_permiso("solo Responsable, Calidad o Admin pueden actualizar no conformidades.")

            tarjeta_seccion(
                "Evidencia",
                "Adjuntar archivo a no conformidad",
                "Permite conservar soportes del hallazgo o del cierre en la misma plataforma.",
            )

            if tiene_rol("responsable", "calidad", "auditor"):
                with st.form("form_evidencia_nc"):
                    evidencia_nc = st.selectbox("No conformidad para evidencia", list(opciones_estado_nc.keys()), key="evidencia_nc")
                    descripcion_evidencia_nc = st.text_input("Descripción de la evidencia", key="descripcion_evidencia_nc")
                    archivo_nc = st.file_uploader(
                        "Archivo de evidencia",
                        key="archivo_evidencia_nc",
                        type=None,
                    )
                    guardar_evidencia_nc = st.form_submit_button("Guardar evidencia")

                if guardar_evidencia_nc:
                    if archivo_nc is None:
                        st.warning("Selecciona un archivo para guardar la evidencia.")
                    else:
                        try:
                            guardar_evidencia_calidad(
                                tipo_entidad="no_conformidad",
                                id_entidad=opciones_estado_nc[evidencia_nc],
                                nombre_archivo_original=archivo_nc.name,
                                contenido_archivo=archivo_nc.getvalue(),
                                descripcion=descripcion_evidencia_nc.strip(),
                                subido_por=st.session_state.get("usuario_email", ""),
                            )
                            st.success("Evidencia guardada correctamente.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"No se pudo guardar la evidencia: {e}")
            else:
                mostrar_aviso_permiso("solo Responsable, Auditor, Calidad o Admin pueden cargar evidencia.")

            evidencia_nc_tabla = []
            for etiqueta, id_nc in opciones_estado_nc.items():
                for id_evidencia, nombre_archivo, descripcion, subido_por, ruta_archivo, created_at in listar_evidencias_calidad("no_conformidad", id_nc):
                    evidencia_nc_tabla.append(
                        {
                            "No conformidad": etiqueta,
                            "Archivo": nombre_archivo,
                            "Descripción": descripcion,
                            "Subido por": subido_por,
                            "Fecha": created_at,
                            "Ruta": ruta_archivo,
                        }
                    )

            if evidencia_nc_tabla:
                st.write("### Evidencias de no conformidades")
                st.dataframe(pd.DataFrame(evidencia_nc_tabla), use_container_width=True, hide_index=True)

            tarjeta_seccion(
                "Aprobación",
                "Cierre formal de no conformidad",
                "El cierre formal deja aprobador, fecha de aprobación y comentario final para auditoría.",
            )

            if tiene_rol("calidad"):
                with st.form("form_cierre_formal_nc"):
                    seleccion_cierre_nc = st.selectbox("No conformidad para cierre formal", list(opciones_estado_nc.keys()), key="seleccion_cierre_formal_nc")
                    verificacion_cierre_formal = st.text_area("Verificación de cierre final", key="verificacion_cierre_formal")
                    comentario_final_nc = st.text_area("Comentario final de aprobación", key="comentario_final_nc")
                    aprobar_cierre_nc = st.form_submit_button("Aprobar cierre formal")

                if aprobar_cierre_nc:
                    try:
                        aprobar_cierre_no_conformidad(
                            id_no_conformidad=opciones_estado_nc[seleccion_cierre_nc],
                            aprobado_por=st.session_state.get("usuario_email", ""),
                            comentario_final=comentario_final_nc.strip(),
                            verificacion_cierre=verificacion_cierre_formal.strip() or None,
                            es_admin=st.session_state.get("es_admin", False),
                            rol_usuario=rol_actual(),
                        )
                        st.success("Cierre formal de no conformidad aprobado.")
                        st.rerun()
                    except PermissionError as e:
                        st.warning(str(e))
                    except Exception as e:
                        st.error(f"No se pudo aprobar el cierre formal: {e}")
            else:
                mostrar_aviso_permiso("solo Calidad o Admin pueden aprobar el cierre formal de no conformidades.")

    with tab2:
        tarjeta_seccion(
            "Acción",
            "Nueva acción correctiva o preventiva",
            "Cada acción queda ligada a una no conformidad para mantener trazabilidad y seguimiento.",
        )

        no_conformidades = listar_no_conformidades()
        opciones_nc = {
            f"{codigo} | {titulo}": id_nc
            for id_nc, codigo, titulo, *_ in no_conformidades
        }

        if not opciones_nc:
            st.info("Primero registra una no conformidad para poder asociar acciones.")
        elif tiene_rol("responsable", "calidad"):
            with st.form("form_accion_calidad", clear_on_submit=True):
                referencia_nc = st.selectbox("No conformidad asociada", list(opciones_nc.keys()))
                col1, col2 = st.columns(2)

                with col1:
                    titulo_accion = st.text_input("Título de la acción")
                    tipo_accion = st.selectbox("Tipo de acción", ["Correctiva", "Preventiva", "Contención"])
                    responsable_accion = st.text_input("Responsable")

                with col2:
                    fecha_inicio = st.date_input("Fecha de inicio")
                    fecha_compromiso_accion = st.date_input("Fecha compromiso", key="fecha_compromiso_accion")
                    descripcion_accion = st.text_area("Descripción de la acción")

                enviada_accion = st.form_submit_button("Guardar acción")

            if enviada_accion:
                datos_accion = {
                    "id_no_conformidad": opciones_nc[referencia_nc],
                    "titulo": titulo_accion.strip(),
                    "descripcion": descripcion_accion.strip(),
                    "tipo_accion": tipo_accion,
                    "responsable": responsable_accion.strip(),
                    "fecha_inicio": str(fecha_inicio),
                    "fecha_compromiso": str(fecha_compromiso_accion),
                    "usuario_email": st.session_state.get("usuario_email", ""),
                }

                campos_requeridos = ["titulo", "descripcion", "responsable"]
                if any(not datos_accion[campo] for campo in campos_requeridos):
                    st.warning("Completa todos los campos obligatorios de la acción.")
                else:
                    try:
                        registrar_accion_calidad(datos_accion)
                        st.success("Acción registrada correctamente.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"No se pudo registrar la acción: {e}")
        else:
            mostrar_aviso_permiso("solo Responsable, Calidad o Admin pueden registrar acciones.")

        acciones = listar_acciones_calidad()
        st.subheader("Acciones registradas")

        if acciones:
            df_acciones = pd.DataFrame(
                acciones,
                columns=[
                    "ID acción", "ID NC", "Código NC", "Título", "Descripción", "Tipo",
                    "Responsable", "Estado", "Fecha inicio", "Fecha compromiso", "Fecha cierre", "Aprobado por", "Fecha aprobación", "Comentario final",
                ],
            )
            colf1, colf2, colf3 = st.columns(3)
            filtro_estado_accion = colf1.selectbox("Filtrar acciones por estado", ["Todos"] + sorted(df_acciones["Estado"].dropna().unique().tolist()), key="filtro_estado_accion")
            filtro_tipo_accion = colf2.selectbox("Filtrar por tipo", ["Todos"] + sorted(df_acciones["Tipo"].dropna().unique().tolist()), key="filtro_tipo_accion")
            filtro_responsable_accion = colf3.selectbox("Filtrar por responsable", ["Todos"] + sorted(df_acciones["Responsable"].dropna().unique().tolist()), key="filtro_responsable_accion")
            df_acciones_filtrado = filtrar_dataframe(
                df_acciones,
                {
                    "Estado": filtro_estado_accion,
                    "Tipo": filtro_tipo_accion,
                    "Responsable": filtro_responsable_accion,
                },
            )
            st.dataframe(df_acciones_filtrado, use_container_width=True, hide_index=True)
        else:
            st.info("Todavía no hay acciones registradas.")

        if acciones and tiene_rol("responsable", "calidad"):
            tarjeta_seccion(
                "Edición",
                "Editar acción",
                "Ajusta título, responsable, fechas o descripción sin perder la trazabilidad en bitácora.",
            )

            acciones_por_label = {
                f"{codigo_nc} | {titulo_accion}": {
                    "id": id_accion,
                    "codigo_nc": codigo_nc,
                    "titulo": titulo_accion,
                    "descripcion": descripcion_accion,
                    "tipo_accion": tipo_accion,
                    "responsable": responsable,
                    "estado": estado,
                    "fecha_inicio": fecha_inicio,
                    "fecha_compromiso": fecha_compromiso,
                    "fecha_cierre": fecha_cierre,
                    "aprobado_por": aprobado_por,
                    "fecha_aprobacion": fecha_aprobacion,
                    "comentario_final": comentario_final,
                }
                for id_accion, _, codigo_nc, titulo_accion, descripcion_accion, tipo_accion, responsable, estado, fecha_inicio, fecha_compromiso, fecha_cierre, aprobado_por, fecha_aprobacion, comentario_final in acciones
            }

            with st.form("form_editar_accion"):
                seleccion_editar_accion = st.selectbox("Acción a editar", list(acciones_por_label.keys()))
                accion_actual = acciones_por_label[seleccion_editar_accion]
                col1, col2 = st.columns(2)
                with col1:
                    editar_titulo_accion = st.text_input("Título", value=accion_actual["titulo"], key="editar_titulo_accion")
                    editar_tipo_accion = st.selectbox("Tipo", ["Correctiva", "Preventiva", "Contención"], index=["Correctiva", "Preventiva", "Contención"].index(accion_actual["tipo_accion"]), key="editar_tipo_accion")
                    editar_responsable_accion = st.text_input("Responsable", value=accion_actual["responsable"], key="editar_responsable_accion")
                with col2:
                    editar_fecha_inicio_accion = st.text_input("Fecha inicio", value=accion_actual["fecha_inicio"], key="editar_fecha_inicio_accion")
                    editar_fecha_compromiso_accion = st.text_input("Fecha compromiso", value=accion_actual["fecha_compromiso"] or "", key="editar_fecha_compromiso_accion")
                    editar_descripcion_accion = st.text_area("Descripción", value=accion_actual["descripcion"], key="editar_descripcion_accion")
                guardar_edicion_accion = st.form_submit_button("Guardar edición de acción")

            if guardar_edicion_accion:
                try:
                    actualizar_accion_calidad(
                        {
                            "id_accion": accion_actual["id"],
                            "titulo": editar_titulo_accion.strip(),
                            "descripcion": editar_descripcion_accion.strip(),
                            "tipo_accion": editar_tipo_accion,
                            "responsable": editar_responsable_accion.strip(),
                            "fecha_inicio": editar_fecha_inicio_accion.strip(),
                            "fecha_compromiso": editar_fecha_compromiso_accion.strip(),
                            "usuario_email": st.session_state.get("usuario_email", ""),
                        }
                    )
                    st.success("Acción editada correctamente.")
                    st.rerun()
                except Exception as e:
                    st.error(f"No se pudo editar la acción: {e}")
        elif acciones:
            mostrar_aviso_permiso("solo Responsable, Calidad o Admin pueden editar acciones.")

            tarjeta_seccion(
                "Estado",
                "Actualizar acción",
                "Las acciones pueden pasar a En proceso o Cerrada y registrar verificación de eficacia.",
            )

            opciones_accion = {
                f"{codigo_nc} | {titulo_accion} | {estado}": id_accion
                for id_accion, _, codigo_nc, titulo_accion, _, _, _, estado, *_ in acciones
            }

            if tiene_rol("responsable", "calidad"):
                with st.form("form_estado_accion"):
                    seleccion_accion = st.selectbox("Selecciona la acción", list(opciones_accion.keys()))
                    nuevo_estado_accion = st.selectbox("Nuevo estado de la acción", ["Abierta", "En proceso"])
                    verificacion_eficacia = st.text_area("Verificación de eficacia", key="verificacion_eficacia")
                    guardar_estado_accion = st.form_submit_button("Actualizar acción")

                if guardar_estado_accion:
                    try:
                        actualizar_estado_accion_calidad(
                            id_accion=opciones_accion[seleccion_accion],
                            nuevo_estado=nuevo_estado_accion,
                            verificacion_eficacia=verificacion_eficacia.strip() or None,
                            es_admin=st.session_state.get("es_admin", False),
                            rol_usuario=rol_actual(),
                            usuario_email=st.session_state.get("usuario_email", ""),
                        )
                        st.success("Acción actualizada correctamente.")
                        st.rerun()
                    except PermissionError as e:
                        st.warning(str(e))
                    except Exception as e:
                        st.error(f"No se pudo actualizar la acción: {e}")
            else:
                mostrar_aviso_permiso("solo Responsable, Calidad o Admin pueden actualizar acciones.")

            tarjeta_seccion(
                "Evidencia",
                "Adjuntar archivo a acción",
                "Útil para guardar planes, formatos, fotos o pruebas de implementación.",
            )

            if tiene_rol("responsable", "calidad", "auditor"):
                with st.form("form_evidencia_accion"):
                    evidencia_accion = st.selectbox("Acción para evidencia", list(opciones_accion.keys()), key="evidencia_accion")
                    descripcion_evidencia_accion = st.text_input("Descripción de la evidencia", key="descripcion_evidencia_accion")
                    archivo_accion = st.file_uploader(
                        "Archivo de evidencia de acción",
                        key="archivo_evidencia_accion",
                        type=None,
                    )
                    guardar_evidencia_accion = st.form_submit_button("Guardar evidencia de acción")

                if guardar_evidencia_accion:
                    if archivo_accion is None:
                        st.warning("Selecciona un archivo para guardar la evidencia.")
                    else:
                        try:
                            guardar_evidencia_calidad(
                                tipo_entidad="accion",
                                id_entidad=opciones_accion[evidencia_accion],
                                nombre_archivo_original=archivo_accion.name,
                                contenido_archivo=archivo_accion.getvalue(),
                                descripcion=descripcion_evidencia_accion.strip(),
                                subido_por=st.session_state.get("usuario_email", ""),
                            )
                            st.success("Evidencia de acción guardada correctamente.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"No se pudo guardar la evidencia: {e}")
            else:
                mostrar_aviso_permiso("solo Responsable, Auditor, Calidad o Admin pueden cargar evidencia.")

            tarjeta_seccion(
                "Aprobación",
                "Cierre formal de acción",
                "El cierre formal deja aprobado por, fecha y comentario final con evidencia de eficacia.",
            )

            if tiene_rol("calidad"):
                with st.form("form_cierre_formal_accion"):
                    seleccion_cierre_accion = st.selectbox("Acción para cierre formal", list(opciones_accion.keys()), key="seleccion_cierre_formal_accion")
                    verificacion_eficacia_formal = st.text_area("Verificación de eficacia final", key="verificacion_eficacia_formal")
                    comentario_final_accion = st.text_area("Comentario final de aprobación", key="comentario_final_accion")
                    aprobar_cierre_accion_btn = st.form_submit_button("Aprobar cierre formal de acción")

                if aprobar_cierre_accion_btn:
                    try:
                        aprobar_cierre_accion(
                            id_accion=opciones_accion[seleccion_cierre_accion],
                            aprobado_por=st.session_state.get("usuario_email", ""),
                            comentario_final=comentario_final_accion.strip(),
                            verificacion_eficacia=verificacion_eficacia_formal.strip() or None,
                            es_admin=st.session_state.get("es_admin", False),
                            rol_usuario=rol_actual(),
                        )
                        st.success("Cierre formal de acción aprobado.")
                        st.rerun()
                    except PermissionError as e:
                        st.warning(str(e))
                    except Exception as e:
                        st.error(f"No se pudo aprobar el cierre formal de la acción: {e}")
            else:
                mostrar_aviso_permiso("solo Calidad o Admin pueden aprobar el cierre formal de acciones.")

    with tab3:
        tarjeta_seccion(
            "Auditoría",
            "Programación de auditorías internas",
            "Deja programadas auditorías, criterios y responsable para fortalecer la preparación de ISO 9001.",
        )

        guardar_auditoria = False
        if tiene_rol("auditor", "calidad"):
            with st.form("form_auditoria", clear_on_submit=True):
                col1, col2 = st.columns(2)

                with col1:
                    codigo_auditoria = st.text_input("Código de auditoría", placeholder="AUD-2026-001")
                    titulo_auditoria = st.text_input("Título")
                    area_auditoria = st.text_input("Área auditada")
                    auditor_lider = st.text_input("Auditor líder")

                with col2:
                    fecha_programada = st.date_input("Fecha programada", key="fecha_programada_auditoria")
                    estado_auditoria = st.selectbox("Estado", ["Programada", "En ejecución", "Cerrada"])
                    alcance_auditoria = st.text_area("Alcance")
                    criterios_auditoria = st.text_area("Criterios")

                guardar_auditoria = st.form_submit_button("Guardar auditoría")

            if guardar_auditoria:
                datos_auditoria = {
                    "codigo": codigo_auditoria.strip(),
                    "titulo": titulo_auditoria.strip(),
                    "area": area_auditoria.strip(),
                    "auditor_lider": auditor_lider.strip(),
                    "fecha_programada": str(fecha_programada),
                    "alcance": alcance_auditoria.strip(),
                    "criterios": criterios_auditoria.strip(),
                    "estado": estado_auditoria,
                    "usuario_email": st.session_state.get("usuario_email", ""),
                }

                campos_requeridos = ["codigo", "titulo", "area", "auditor_lider"]
                if any(not datos_auditoria[campo] for campo in campos_requeridos):
                    st.warning("Completa los campos obligatorios de la auditoría.")
                else:
                    try:
                        registrar_auditoria_calidad(datos_auditoria)
                        st.success("Auditoría registrada correctamente.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"No se pudo registrar la auditoría: {e}")
        else:
            mostrar_aviso_permiso("solo Auditor, Calidad o Admin pueden programar auditorías.")

        auditorias = listar_auditorias_calidad()
        st.subheader("Auditorías internas")

        if auditorias:
            df_auditorias = pd.DataFrame(
                auditorias,
                columns=["ID", "Código", "Título", "Área", "Auditor líder", "Fecha programada", "Estado", "Resultado"],
            )
            colf1, colf2 = st.columns(2)
            filtro_estado_auditoria = colf1.selectbox("Filtrar auditorías por estado", ["Todos"] + sorted(df_auditorias["Estado"].dropna().unique().tolist()), key="filtro_estado_auditoria")
            filtro_area_auditoria = colf2.selectbox("Filtrar auditorías por área", ["Todos"] + sorted(df_auditorias["Área"].dropna().unique().tolist()), key="filtro_area_auditoria")
            df_auditorias_filtrado = filtrar_dataframe(
                df_auditorias,
                {
                    "Estado": filtro_estado_auditoria,
                    "Área": filtro_area_auditoria,
                },
            )
            st.dataframe(df_auditorias_filtrado, use_container_width=True, hide_index=True)
        else:
            st.info("Todavía no hay auditorías registradas.")

        tarjeta_seccion(
            "Hallazgo",
            "Registrar hallazgo de auditoría",
            "Cada auditoría puede generar hallazgos con severidad, responsable y fecha compromiso.",
        )

        if auditorias:
            opciones_auditoria = {
                f"{codigo} | {titulo}": id_auditoria
                for id_auditoria, codigo, titulo, *_ in auditorias
            }

            if tiene_rol("auditor", "calidad"):
                with st.form("form_hallazgo_auditoria", clear_on_submit=True):
                    auditoria_hallazgo = st.selectbox("Auditoría asociada", list(opciones_auditoria.keys()))
                    referencia_hallazgo = st.text_input("Referencia", placeholder="ISO 9001 - 8.7")
                    descripcion_hallazgo = st.text_area("Descripción del hallazgo")
                    col1, col2 = st.columns(2)
                    with col1:
                        severidad_hallazgo = st.selectbox("Severidad del hallazgo", ["Menor", "Mayor", "Crítica"])
                        responsable_hallazgo = st.text_input("Responsable")
                    with col2:
                        estado_hallazgo = st.selectbox("Estado del hallazgo", ["Abierto", "En proceso", "Cerrado"])
                        fecha_compromiso_hallazgo = st.date_input("Fecha compromiso", key="fecha_compromiso_hallazgo")
                    guardar_hallazgo = st.form_submit_button("Guardar hallazgo")

                if guardar_hallazgo:
                    datos_hallazgo = {
                        "id_auditoria": opciones_auditoria[auditoria_hallazgo],
                        "referencia": referencia_hallazgo.strip(),
                        "descripcion": descripcion_hallazgo.strip(),
                        "severidad": severidad_hallazgo,
                        "estado": estado_hallazgo,
                        "responsable": responsable_hallazgo.strip(),
                        "fecha_compromiso": str(fecha_compromiso_hallazgo),
                        "usuario_email": st.session_state.get("usuario_email", ""),
                    }

                    if not datos_hallazgo["referencia"] or not datos_hallazgo["descripcion"]:
                        st.warning("Completa la referencia y la descripción del hallazgo.")
                    else:
                        try:
                            registrar_hallazgo_auditoria(datos_hallazgo)
                            st.success("Hallazgo registrado correctamente.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"No se pudo registrar el hallazgo: {e}")
            else:
                mostrar_aviso_permiso("solo Auditor, Calidad o Admin pueden registrar hallazgos de auditoría.")

        hallazgos = listar_hallazgos_auditoria()
        if hallazgos:
            st.write("### Hallazgos de auditoría")
            df_hallazgos = pd.DataFrame(
                hallazgos,
                columns=["ID hallazgo", "ID auditoría", "Código auditoría", "Referencia", "Descripción", "Severidad", "Estado", "Responsable", "Fecha compromiso"],
            )
            st.dataframe(df_hallazgos, use_container_width=True, hide_index=True)

    with tab4:
        tarjeta_seccion(
            "Documento",
            "Control documental",
            "Gestiona documentos, versiones, vigencia y aprobaciones dentro del sistema de calidad.",
        )

        guardar_documento = False
        if tiene_rol("calidad"):
            with st.form("form_documento", clear_on_submit=True):
                col1, col2 = st.columns(2)
                with col1:
                    codigo_documento = st.text_input("Código de documento", placeholder="PROC-CAL-001")
                    nombre_documento = st.text_input("Nombre del documento")
                    proceso_documento = st.text_input("Proceso o área")
                with col2:
                    tipo_documento = st.selectbox("Tipo de documento", ["Procedimiento", "Formato", "Instructivo", "Política", "Manual", "Registro"])
                    estado_documento = st.selectbox("Estado inicial", ["Borrador", "Vigente", "Obsoleto"])
                    observaciones_documento = st.text_area("Observaciones")
                guardar_documento = st.form_submit_button("Guardar documento")

            if guardar_documento:
                datos_documento = {
                    "codigo": codigo_documento.strip(),
                    "nombre": nombre_documento.strip(),
                    "proceso_area": proceso_documento.strip(),
                    "tipo_documento": tipo_documento,
                    "estado": estado_documento,
                    "observaciones": observaciones_documento.strip(),
                    "usuario_email": st.session_state.get("usuario_email", ""),
                }
                if not datos_documento["codigo"] or not datos_documento["nombre"] or not datos_documento["proceso_area"]:
                    st.warning("Completa código, nombre y proceso o área del documento.")
                else:
                    try:
                        registrar_documento_calidad(datos_documento)
                        st.success("Documento registrado correctamente.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"No se pudo registrar el documento: {e}")
        else:
            mostrar_aviso_permiso("solo Calidad o Admin pueden crear documentos controlados.")

        documentos = listar_documentos_calidad()
        st.subheader("Documentos controlados")

        if documentos:
            df_documentos = pd.DataFrame(
                documentos,
                columns=[
                    "ID", "Código", "Nombre", "Proceso/Área", "Tipo",
                    "Estado", "Versión actual", "Vigente desde", "Vigente hasta",
                    "Aprobado por", "Fecha aprobación", "Observaciones",
                ],
            )
            colf1, colf2, colf3 = st.columns(3)
            filtro_estado_doc = colf1.selectbox("Filtrar documentos por estado", ["Todos"] + sorted(df_documentos["Estado"].dropna().unique().tolist()), key="filtro_estado_doc")
            filtro_tipo_doc = colf2.selectbox("Filtrar por tipo", ["Todos"] + sorted(df_documentos["Tipo"].dropna().unique().tolist()), key="filtro_tipo_doc")
            filtro_area_doc = colf3.selectbox("Filtrar por área", ["Todos"] + sorted(df_documentos["Proceso/Área"].dropna().unique().tolist()), key="filtro_area_doc")
            df_documentos_filtrado = filtrar_dataframe(
                df_documentos,
                {
                    "Estado": filtro_estado_doc,
                    "Tipo": filtro_tipo_doc,
                    "Proceso/Área": filtro_area_doc,
                },
            )
            st.dataframe(df_documentos_filtrado, use_container_width=True, hide_index=True)
        else:
            st.info("Todavía no hay documentos controlados.")

        if documentos:
            documentos_por_label = {
                f"{codigo} | {nombre}": id_documento
                for id_documento, codigo, nombre, *_ in documentos
            }

            tarjeta_seccion(
                "Versión",
                "Registrar nueva versión documental",
                "Sube una nueva versión, describe los cambios y conserva el archivo asociado.",
            )

            if tiene_rol("calidad"):
                with st.form("form_version_documento", clear_on_submit=True):
                    seleccion_documento_version = st.selectbox("Documento", list(documentos_por_label.keys()))
                    version_documento = st.text_input("Versión", placeholder="1.0")
                    cambios_version = st.text_area("Resumen de cambios")
                    archivo_version = st.file_uploader("Archivo del documento", key="archivo_documento_version")
                    guardar_version_documento = st.form_submit_button("Guardar versión")

                if guardar_version_documento:
                    try:
                        registrar_version_documento(
                            id_documento=documentos_por_label[seleccion_documento_version],
                            version=version_documento.strip(),
                            cambios_resumen=cambios_version.strip(),
                            elaborado_por=st.session_state.get("usuario_email", ""),
                            nombre_archivo_original=archivo_version.name if archivo_version else None,
                            contenido_archivo=archivo_version.getvalue() if archivo_version else None,
                        )
                        st.success("Versión documental registrada correctamente.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"No se pudo registrar la versión: {e}")
            else:
                mostrar_aviso_permiso("solo Calidad o Admin pueden registrar versiones documentales.")

            tarjeta_seccion(
                "Aprobación",
                "Aprobar versión y vigencia",
                "La aprobación formal marca la versión vigente y actualiza el estado del documento.",
            )

            if tiene_rol("calidad"):
                with st.form("form_aprobar_version_documento"):
                    seleccion_documento_aprobacion = st.selectbox("Documento para aprobar", list(documentos_por_label.keys()), key="seleccion_documento_aprobacion")
                    id_doc_aprobar = documentos_por_label[seleccion_documento_aprobacion]
                    versiones_documento = listar_versiones_documento(id_doc_aprobar)
                    opciones_version = {
                        f"{version} | {'Vigente' if es_vigente else 'Pendiente'}": id_version
                        for id_version, version, _, _, _, _, _, _, es_vigente, _ in versiones_documento
                    }
                    seleccion_version = st.selectbox("Versión", list(opciones_version.keys())) if opciones_version else st.selectbox("Versión", ["Sin versiones"])
                    col1, col2 = st.columns(2)
                    with col1:
                        vigente_desde_doc = st.date_input("Vigente desde", key="vigente_desde_doc")
                    with col2:
                        vigente_hasta_doc = st.date_input("Vigente hasta", key="vigente_hasta_doc")
                    estado_aprobado_doc = st.selectbox("Estado del documento", ["Vigente", "Obsoleto"], key="estado_aprobado_doc")
                    aprobar_version_btn = st.form_submit_button("Aprobar versión documental")

                if aprobar_version_btn:
                    if not opciones_version:
                        st.warning("Primero registra una versión para poder aprobarla.")
                    else:
                        try:
                            aprobar_version_documento(
                                id_documento=id_doc_aprobar,
                                id_version=opciones_version[seleccion_version],
                                aprobado_por=st.session_state.get("usuario_email", ""),
                                vigente_desde=str(vigente_desde_doc),
                                vigente_hasta=str(vigente_hasta_doc),
                                estado_documento=estado_aprobado_doc,
                                es_admin=st.session_state.get("es_admin", False),
                                rol_usuario=rol_actual(),
                            )
                            st.success("Versión documental aprobada correctamente.")
                            st.rerun()
                        except PermissionError as e:
                            st.warning(str(e))
                        except Exception as e:
                            st.error(f"No se pudo aprobar la versión documental: {e}")
            else:
                mostrar_aviso_permiso("solo Calidad o Admin pueden aprobar versiones documentales.")

            st.write("### Versiones documentales")
            tabla_versiones = []
            for etiqueta, id_documento in documentos_por_label.items():
                for id_version, version, nombre_archivo, ruta_archivo, cambios_resumen, elaborado_por, aprobado_por, fecha_aprobacion, es_vigente, created_at in listar_versiones_documento(id_documento):
                    tabla_versiones.append(
                        {
                            "Documento": etiqueta,
                            "Versión": version,
                            "Archivo": nombre_archivo,
                            "Cambios": cambios_resumen,
                            "Elaborado por": elaborado_por,
                            "Aprobado por": aprobado_por,
                            "Fecha aprobación": fecha_aprobacion,
                            "Vigente": "Sí" if es_vigente else "No",
                            "Fecha registro": created_at,
                        }
                    )

            if tabla_versiones:
                st.dataframe(pd.DataFrame(tabla_versiones), use_container_width=True, hide_index=True)

                st.write("### Descarga de versiones")
                for etiqueta, id_documento in documentos_por_label.items():
                    for id_version, version, nombre_archivo, ruta_archivo, _, _, _, _, es_vigente, _ in listar_versiones_documento(id_documento):
                        if nombre_archivo and ruta_archivo:
                            try:
                                with open(ruta_archivo, "rb") as archivo:
                                    st.download_button(
                                        label=f"Descargar {etiqueta} v{version}{' (vigente)' if es_vigente else ''}",
                                        data=archivo.read(),
                                        file_name=nombre_archivo,
                                        key=f"descarga_documento_{id_version}",
                                    )
                            except FileNotFoundError:
                                st.warning(f"No se encontró el archivo de la versión {version} para {etiqueta}.")

    with tab5:
        tarjeta_seccion(
            "Control",
            "Vista de seguimiento",
            "Resumen rápido para revisar carga operativa, prioridades y trazabilidad del sistema de calidad.",
        )

        no_conformidades = listar_no_conformidades()
        acciones = listar_acciones_calidad()
        auditorias = listar_auditorias_calidad()
        hallazgos = listar_hallazgos_auditoria()
        documentos = listar_documentos_calidad()
        df_nc_exec = pd.DataFrame()
        df_acc_exec = pd.DataFrame()
        df_aud_exec = pd.DataFrame()
        df_hall_exec = pd.DataFrame()
        df_docs_exec = pd.DataFrame()
        df_versiones_exec = pd.DataFrame()

        if no_conformidades:
            df_nc_exec = pd.DataFrame(
                no_conformidades,
                columns=[
                    "ID", "Código", "Título", "Descripción", "Origen", "Área", "Severidad",
                    "Estado", "Detectado por", "Responsable", "Fecha detección",
                    "Fecha compromiso", "Causa raíz", "Fecha cierre", "Aprobado por", "Fecha aprobación", "Comentario final",
                ],
            )
            df_nc_exec["Semáforo"] = df_nc_exec.apply(
                lambda fila: semaforo_no_conformidad(fila["Estado"], fila["Severidad"]),
                axis=1,
            )

        if acciones:
            df_acc_exec = pd.DataFrame(
                acciones,
                columns=[
                    "ID acción", "ID NC", "Código NC", "Título", "Descripción", "Tipo",
                    "Responsable", "Estado", "Fecha inicio", "Fecha compromiso", "Fecha cierre", "Aprobado por", "Fecha aprobación", "Comentario final",
                ],
            )
            df_acc_exec["Semáforo"] = df_acc_exec.apply(
                lambda fila: semaforo_accion(fila["Estado"], fila["Fecha compromiso"]),
                axis=1,
            )
            df_acc_exec["Fecha compromiso"] = pd.to_datetime(df_acc_exec["Fecha compromiso"], errors="coerce")

        if auditorias:
            df_aud_exec = pd.DataFrame(
                auditorias,
                columns=[
                    "ID auditoría", "Código", "Título", "Área", "Auditor líder",
                    "Fecha programada", "Estado", "Resultado",
                ],
            )

        if hallazgos:
            df_hall_exec = pd.DataFrame(
                hallazgos,
                columns=[
                    "ID hallazgo", "ID auditoría", "Código auditoría", "Referencia",
                    "Descripción", "Severidad", "Estado", "Responsable", "Fecha compromiso",
                ],
            )

        if documentos:
            df_docs_exec = pd.DataFrame(
                documentos,
                columns=[
                    "ID", "Código", "Nombre", "Proceso/Área", "Tipo",
                    "Estado", "Versión actual", "Vigente desde", "Vigente hasta",
                    "Aprobado por", "Fecha aprobación", "Observaciones",
                ],
            )

            versiones_tablero = []
            for id_documento, codigo, nombre, *_ in documentos:
                for id_version, version, _, _, _, elaborado_por, aprobado_por, fecha_aprobacion, es_vigente, created_at in listar_versiones_documento(id_documento):
                    versiones_tablero.append(
                        {
                            "ID documento": id_documento,
                            "Código": codigo,
                            "Nombre": nombre,
                            "ID versión": id_version,
                            "Versión": version,
                            "Elaborado por": elaborado_por,
                            "Aprobado por": aprobado_por,
                            "Fecha aprobación": fecha_aprobacion,
                            "Vigente": "Sí" if es_vigente else "No",
                            "Registrada el": created_at,
                        }
                    )

            if versiones_tablero:
                df_versiones_exec = pd.DataFrame(versiones_tablero)

        st.write("### Alertas operativas")
        alerta_col1, alerta_col2, alerta_col3, alerta_col4 = st.columns(4)

        acciones_vencidas = (
            df_acc_exec[
                (df_acc_exec["Estado"] != "Cerrada")
                & (df_acc_exec["Fecha compromiso"].notna())
                & (df_acc_exec["Fecha compromiso"] < pd.Timestamp.today().normalize())
            ]
            if not df_acc_exec.empty
            else pd.DataFrame()
        )
        nc_criticas = (
            df_nc_exec[
                (df_nc_exec["Estado"] != "Cerrada")
                & (df_nc_exec["Severidad"].isin(["Alta", "Crítica"]))
            ]
            if not df_nc_exec.empty
            else pd.DataFrame()
        )
        auditorias_pendientes = (
            len([fila for fila in auditorias if fila[6] != "Cerrada"])
        )
        documentos_por_vencer = (
            df_docs_exec[
                (df_docs_exec["Estado"] == "Vigente")
                & (pd.to_datetime(df_docs_exec["Vigente hasta"], errors="coerce").notna())
                & (
                    pd.to_datetime(df_docs_exec["Vigente hasta"], errors="coerce")
                    <= (pd.Timestamp.today().normalize() + pd.Timedelta(days=30))
                )
            ]
            if not df_docs_exec.empty
            else pd.DataFrame()
        )
        recordatorios_calidad = construir_recordatorios_calidad(
            df_nc_exec,
            df_acc_exec,
            df_aud_exec,
            df_docs_exec,
            df_hall_exec,
            df_versiones_exec,
        )
        pendientes_prioritarios = construir_pendientes_prioritarios(
            df_nc_exec,
            df_acc_exec,
            df_docs_exec,
            df_aud_exec,
            df_hall_exec,
            df_versiones_exec,
        )

        with alerta_col1:
            if not acciones_vencidas.empty:
                st.error(f"{len(acciones_vencidas)} acciones vencidas requieren atención.")
            else:
                st.success("No hay acciones vencidas.")

        with alerta_col2:
            if not nc_criticas.empty:
                st.warning(f"{len(nc_criticas)} no conformidades altas o críticas siguen abiertas.")
            else:
                st.success("No hay no conformidades críticas abiertas.")

        with alerta_col3:
            if auditorias_pendientes > 0:
                st.info(f"{auditorias_pendientes} auditorías siguen programadas o en ejecución.")
            else:
                st.success("No hay auditorías pendientes.")

        with alerta_col4:
            if not documentos_por_vencer.empty:
                st.warning(f"{len(documentos_por_vencer)} documentos están por vencer en los próximos 30 días.")
            else:
                st.success("No hay documentos próximos a vencerse.")

        tarjeta_seccion(
            "Prioridad",
            "Panel de pendientes prioritarios",
            "Consolida cierres pendientes, vencimientos, hallazgos críticos y aprobaciones documentales en una sola vista.",
        )

        if not pendientes_prioritarios.empty:
            prioridad_col1, prioridad_col2, prioridad_col3 = st.columns(3)
            with prioridad_col1:
                st.metric("Pendientes prioritarios", len(pendientes_prioritarios))
            with prioridad_col2:
                st.metric(
                    "Pendientes críticos",
                    len(pendientes_prioritarios[pendientes_prioritarios["Prioridad"] == "Crítica"]),
                )
            with prioridad_col3:
                st.metric(
                    "Pendientes altos",
                    len(pendientes_prioritarios[pendientes_prioritarios["Prioridad"] == "Alta"]),
                )

            filtro_prioridad = st.multiselect(
                "Filtrar panel por prioridad",
                ["Crítica", "Alta", "Media", "Baja"],
                default=["Crítica", "Alta", "Media"],
                key="filtro_prioridad_pendientes",
            )
            pendientes_visibles = pendientes_prioritarios[
                pendientes_prioritarios["Prioridad"].isin(filtro_prioridad)
            ] if filtro_prioridad else pendientes_prioritarios
            st.dataframe(pendientes_visibles, use_container_width=True, hide_index=True)
        else:
            pendientes_visibles = pendientes_prioritarios
            st.success("No hay pendientes prioritarios acumulados en este momento.")

        tarjeta_seccion(
            "Fechas",
            "Recordatorios por fechas críticas",
            "Resume próximos vencimientos y elementos fuera de plazo para que el equipo actúe antes del cierre.",
        )

        if not recordatorios_calidad.empty:
            recordatorio_col1, recordatorio_col2 = st.columns(2)
            with recordatorio_col1:
                filtro_tipo_recordatorio = st.multiselect(
                    "Tipos de recordatorio",
                    sorted(recordatorios_calidad["Tipo"].dropna().unique().tolist()),
                    default=sorted(recordatorios_calidad["Tipo"].dropna().unique().tolist()),
                    key="filtro_tipo_recordatorio",
                )
            with recordatorio_col2:
                filtro_alerta_recordatorio = st.multiselect(
                    "Estado de alerta",
                    ["Vencido", "Vence hoy", "Vence pronto"],
                    default=["Vencido", "Vence hoy", "Vence pronto"],
                    key="filtro_alerta_recordatorio",
                )

            recordatorios_visibles = recordatorios_calidad.copy()
            if filtro_tipo_recordatorio:
                recordatorios_visibles = recordatorios_visibles[
                    recordatorios_visibles["Tipo"].isin(filtro_tipo_recordatorio)
                ]
            if filtro_alerta_recordatorio:
                recordatorios_visibles = recordatorios_visibles[
                    recordatorios_visibles["Alerta"].isin(filtro_alerta_recordatorio)
                ]

            col_recordatorio_1, col_recordatorio_2, col_recordatorio_3 = st.columns(3)
            with col_recordatorio_1:
                st.metric("Recordatorios activos", len(recordatorios_visibles))
            with col_recordatorio_2:
                st.metric(
                    "Vencidos",
                    len(recordatorios_visibles[recordatorios_visibles["Alerta"] == "Vencido"]),
                )
            with col_recordatorio_3:
                st.metric(
                    "Vence hoy",
                    len(recordatorios_visibles[recordatorios_visibles["Alerta"] == "Vence hoy"]),
                )

            st.dataframe(recordatorios_visibles, use_container_width=True, hide_index=True)
        else:
            recordatorios_visibles = recordatorios_calidad
            st.success("No hay recordatorios activos por fechas críticas.")

        tarjeta_seccion(
            "Filtros",
            "Filtro ejecutivo",
            "Usa estos filtros para concentrarte en prioridades y exportar exactamente la vista que necesitas.",
        )

        exec_filter_col1, exec_filter_col2, exec_filter_col3, exec_filter_col4 = st.columns(4)
        filtro_exec_estado_nc = exec_filter_col1.selectbox(
            "Estado NC",
            ["Todos"] + (sorted(df_nc_exec["Estado"].dropna().unique().tolist()) if not df_nc_exec.empty else []),
            key="filtro_exec_estado_nc",
        )
        filtro_exec_area_nc = exec_filter_col2.selectbox(
            "Área NC",
            ["Todos"] + (sorted(df_nc_exec["Área"].dropna().unique().tolist()) if not df_nc_exec.empty else []),
            key="filtro_exec_area_nc",
        )
        filtro_exec_estado_acc = exec_filter_col3.selectbox(
            "Estado acción",
            ["Todos"] + (sorted(df_acc_exec["Estado"].dropna().unique().tolist()) if not df_acc_exec.empty else []),
            key="filtro_exec_estado_acc",
        )
        filtro_exec_resp_acc = exec_filter_col4.selectbox(
            "Responsable acción",
            ["Todos"] + (sorted(df_acc_exec["Responsable"].dropna().unique().tolist()) if not df_acc_exec.empty else []),
            key="filtro_exec_resp_acc",
        )

        if not df_nc_exec.empty:
            df_nc_exec_filtrado = filtrar_dataframe(
                df_nc_exec,
                {
                    "Estado": filtro_exec_estado_nc,
                    "Área": filtro_exec_area_nc,
                },
            )
        else:
            df_nc_exec_filtrado = df_nc_exec

        if not df_acc_exec.empty:
            df_acc_exec_filtrado = filtrar_dataframe(
                df_acc_exec,
                {
                    "Estado": filtro_exec_estado_acc,
                    "Responsable": filtro_exec_resp_acc,
                },
            )
        else:
            df_acc_exec_filtrado = df_acc_exec

        col1, col2 = st.columns(2)

        with col1:
            if not df_nc_exec_filtrado.empty:
                st.write("### Distribución por severidad")
                st.bar_chart(df_nc_exec_filtrado["Severidad"].value_counts())
            else:
                st.info("Sin datos de no conformidades para graficar.")

        with col2:
            if not df_acc_exec_filtrado.empty:
                st.write("### Distribución por estado de acciones")
                st.bar_chart(df_acc_exec_filtrado["Estado"].value_counts())
            else:
                st.info("Sin datos de acciones para graficar.")

        st.write("### Tablero ejecutivo")
        exec_col1, exec_col2 = st.columns(2)

        with exec_col1:
            if not df_nc_exec_filtrado.empty:
                st.write("#### No conformidades por área")
                st.bar_chart(df_nc_exec_filtrado["Área"].value_counts())
            else:
                st.info("Sin no conformidades para el tablero.")

        with exec_col2:
            if not df_acc_exec_filtrado.empty:
                st.write("#### Acciones por responsable")
                st.bar_chart(df_acc_exec_filtrado["Responsable"].value_counts())
            else:
                st.info("Sin acciones para el tablero.")

        if not df_acc_exec_filtrado.empty:
            vencidas = df_acc_exec_filtrado[
                (df_acc_exec_filtrado["Estado"] != "Cerrada")
                & (df_acc_exec_filtrado["Fecha compromiso"].notna())
                & (df_acc_exec_filtrado["Fecha compromiso"] < pd.Timestamp.today().normalize())
            ]
            st.write("#### Acciones vencidas prioritarias")
            if not vencidas.empty:
                st.dataframe(dataframe_con_semaforo(vencidas), use_container_width=True, hide_index=True)
            else:
                st.info("No hay acciones vencidas en este momento.")
        else:
            vencidas = pd.DataFrame()

        if not df_nc_exec_filtrado.empty:
            st.write("#### Hallazgos abiertos críticos o altos")
            prioridades_nc = df_nc_exec_filtrado[
                (df_nc_exec_filtrado["Estado"] != "Cerrada")
                & (df_nc_exec_filtrado["Severidad"].isin(["Alta", "Crítica"]))
            ]
            if not prioridades_nc.empty:
                st.dataframe(dataframe_con_semaforo(prioridades_nc), use_container_width=True, hide_index=True)
            else:
                st.info("No hay no conformidades altas o críticas abiertas.")
        else:
            prioridades_nc = pd.DataFrame()

        if not df_nc_exec_filtrado.empty:
            st.write("#### No conformidades filtradas")
            st.dataframe(dataframe_con_semaforo(df_nc_exec_filtrado), use_container_width=True, hide_index=True)

        if not df_acc_exec_filtrado.empty:
            st.write("#### Acciones filtradas")
            st.dataframe(dataframe_con_semaforo(df_acc_exec_filtrado), use_container_width=True, hide_index=True)

        resumen_lineas = [
            f"No conformidades visibles en tablero: {len(df_nc_exec_filtrado)}",
            f"Acciones visibles en tablero: {len(df_acc_exec_filtrado)}",
            f"Acciones vencidas visibles: {len(acciones_vencidas)}",
            f"No conformidades altas o críticas visibles: {len(prioridades_nc)}",
            f"Pendientes prioritarios activos: {len(pendientes_visibles)}",
            f"Recordatorios por fechas críticas activos: {len(recordatorios_visibles)}",
        ]

        secciones_exportables = {
            "No conformidades": df_nc_exec_filtrado,
            "Acciones": df_acc_exec_filtrado,
            "Acciones vencidas": acciones_vencidas,
            "Prioridades NC": prioridades_nc,
            "Pendientes prioritarios": pendientes_visibles,
            "Recordatorios críticos": recordatorios_visibles,
        }

        export_col1, export_col2 = st.columns(2)
        with export_col1:
            st.download_button(
                label="Descargar reporte ejecutivo en Excel",
                data=generar_excel_ejecutivo(secciones_exportables),
                file_name="reporte_ejecutivo_calidad.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with export_col2:
            try:
                pdf_bytes = generar_pdf_ejecutivo(resumen_lineas, secciones_exportables)
                st.download_button(
                    label="Descargar reporte ejecutivo en PDF",
                    data=pdf_bytes,
                    file_name="reporte_ejecutivo_calidad.pdf",
                    mime="application/pdf",
                )
            except RuntimeError as exc:
                st.warning(str(exc))

        evidencia_total = []
        for id_nc, *_ in no_conformidades:
            evidencia_total.extend(listar_evidencias_calidad("no_conformidad", id_nc))
        for id_accion, *_ in acciones:
            evidencia_total.extend(listar_evidencias_calidad("accion", id_accion))

        if evidencia_total:
            st.write("### Descarga de evidencias")
            for id_evidencia, nombre_archivo, descripcion, subido_por, ruta_archivo, created_at in evidencia_total:
                try:
                    with open(ruta_archivo, "rb") as archivo:
                        st.download_button(
                            label=f"Descargar {nombre_archivo}",
                            data=archivo.read(),
                            file_name=nombre_archivo,
                            key=f"descarga_evidencia_{id_evidencia}",
                        )
                    if descripcion:
                        st.caption(f"{descripcion} | {subido_por} | {created_at}")
                except FileNotFoundError:
                    st.warning(f"No se encontró el archivo {nombre_archivo} en la ruta almacenada.")

        tarjeta_seccion(
            "Bitácora",
            "Trazabilidad de cambios",
            "Registro cronológico de altas, cambios de estado, evidencias y eventos de auditoría.",
        )

        bitacora = listar_bitacora_calidad()
        if bitacora:
            df_bitacora = pd.DataFrame(
                bitacora,
                columns=["ID", "Entidad", "ID entidad", "Acción", "Detalle", "Usuario", "Fecha"],
            )
            st.dataframe(df_bitacora, use_container_width=True, hide_index=True)
        else:
            st.info("Todavía no hay eventos registrados en la bitácora.")

    with tab6:
        tarjeta_seccion(
            "KPI",
            "Indicadores de calidad",
            "Mide cumplimiento, tiempos de cierre y carga operativa para dar seguimiento ejecutivo al sistema.",
        )

        meses_analisis = st.selectbox(
            "Ventana de análisis",
            [3, 6, 12],
            index=1,
            format_func=lambda valor: f"Últimos {valor} meses",
            key="indicadores_meses",
        )

        nc_ind = listar_no_conformidades()
        acciones_ind = listar_acciones_calidad()
        auditorias_ind = listar_auditorias_calidad()
        hallazgos_ind = listar_hallazgos_auditoria()
        documentos_ind = listar_documentos_calidad()

        df_nc_ind = pd.DataFrame(
            nc_ind,
            columns=[
                "ID", "Código", "Título", "Descripción", "Origen", "Área", "Severidad",
                "Estado", "Detectado por", "Responsable", "Fecha detección",
                "Fecha compromiso", "Causa raíz", "Fecha cierre", "Aprobado por", "Fecha aprobación", "Comentario final",
            ],
        ) if nc_ind else pd.DataFrame()
        df_acc_ind = pd.DataFrame(
            acciones_ind,
            columns=[
                "ID acción", "ID NC", "Código NC", "Título", "Descripción", "Tipo",
                "Responsable", "Estado", "Fecha inicio", "Fecha compromiso", "Fecha cierre", "Aprobado por", "Fecha aprobación", "Comentario final",
            ],
        ) if acciones_ind else pd.DataFrame()
        df_aud_ind = pd.DataFrame(
            auditorias_ind,
            columns=[
                "ID auditoría", "Código", "Título", "Área", "Auditor líder",
                "Fecha programada", "Estado", "Resultado",
            ],
        ) if auditorias_ind else pd.DataFrame()
        df_hall_ind = pd.DataFrame(
            hallazgos_ind,
            columns=[
                "ID hallazgo", "ID auditoría", "Código auditoría", "Referencia",
                "Descripción", "Severidad", "Estado", "Responsable", "Fecha compromiso",
            ],
        ) if hallazgos_ind else pd.DataFrame()
        df_docs_ind = pd.DataFrame(
            documentos_ind,
            columns=[
                "ID", "Código", "Nombre", "Proceso/Área", "Tipo",
                "Estado", "Versión actual", "Vigente desde", "Vigente hasta",
                "Aprobado por", "Fecha aprobación", "Observaciones",
            ],
        ) if documentos_ind else pd.DataFrame()

        versiones_ind = []
        for id_documento, codigo, nombre, *_ in documentos_ind:
            for id_version, version, _, _, _, elaborado_por, aprobado_por, fecha_aprobacion, es_vigente, created_at in listar_versiones_documento(id_documento):
                versiones_ind.append(
                    {
                        "Código": codigo,
                        "Nombre": nombre,
                        "Versión": version,
                        "Elaborado por": elaborado_por,
                        "Aprobado por": aprobado_por,
                        "Fecha aprobación": fecha_aprobacion,
                        "Vigente": "Sí" if es_vigente else "No",
                        "Registrada el": created_at,
                    }
                )
        df_versiones_ind = pd.DataFrame(versiones_ind) if versiones_ind else pd.DataFrame()

        total_nc = len(df_nc_ind)
        nc_abiertas = len(df_nc_ind[df_nc_ind["Estado"] != "Cerrada"]) if not df_nc_ind.empty else 0
        nc_cerradas = len(df_nc_ind[df_nc_ind["Estado"] == "Cerrada"]) if not df_nc_ind.empty else 0
        total_acciones = len(df_acc_ind)
        acciones_abiertas = len(df_acc_ind[df_acc_ind["Estado"] != "Cerrada"]) if not df_acc_ind.empty else 0
        acciones_cerradas = len(df_acc_ind[df_acc_ind["Estado"] == "Cerrada"]) if not df_acc_ind.empty else 0
        acciones_vencidas_ind = len(
            df_acc_ind[
                (df_acc_ind["Estado"] != "Cerrada")
                & (pd.to_datetime(df_acc_ind["Fecha compromiso"], errors="coerce") < pd.Timestamp.today().normalize())
            ]
        ) if not df_acc_ind.empty else 0
        auditorias_cerradas = len(df_aud_ind[df_aud_ind["Estado"] == "Cerrada"]) if not df_aud_ind.empty else 0
        total_auditorias = len(df_aud_ind)
        documentos_vigentes = len(df_docs_ind[df_docs_ind["Estado"] == "Vigente"]) if not df_docs_ind.empty else 0
        documentos_vencidos = len(
            df_docs_ind[
                (df_docs_ind["Estado"] == "Vigente")
                & (pd.to_datetime(df_docs_ind["Vigente hasta"], errors="coerce") < pd.Timestamp.today().normalize())
            ]
        ) if not df_docs_ind.empty else 0
        versiones_pendientes = len(
            df_versiones_ind[
                df_versiones_ind["Aprobado por"].isna() | (df_versiones_ind["Aprobado por"].astype(str).str.strip() == "")
            ]
        ) if not df_versiones_ind.empty else 0

        promedio_cierre_nc = calcular_promedio_cierre(df_nc_ind, "Fecha detección", "Fecha cierre")
        promedio_cierre_acciones = calcular_promedio_cierre(df_acc_ind, "Fecha inicio", "Fecha cierre")

        kpi_col1, kpi_col2, kpi_col3, kpi_col4 = st.columns(4)
        with kpi_col1:
            st.metric("NC abiertas", nc_abiertas, delta=f"Cierre: {formatear_porcentaje(nc_cerradas, total_nc)}")
        with kpi_col2:
            st.metric("Acciones vencidas", acciones_vencidas_ind, delta=f"Abiertas: {acciones_abiertas}")
        with kpi_col3:
            st.metric("Promedio cierre NC", f"{promedio_cierre_nc} días")
        with kpi_col4:
            st.metric("Cumplimiento auditorías", formatear_porcentaje(auditorias_cerradas, total_auditorias))

        kpi_col5, kpi_col6, kpi_col7, kpi_col8 = st.columns(4)
        with kpi_col5:
            st.metric("Acciones cerradas", acciones_cerradas, delta=formatear_porcentaje(acciones_cerradas, total_acciones))
        with kpi_col6:
            st.metric("Promedio cierre acciones", f"{promedio_cierre_acciones} días")
        with kpi_col7:
            st.metric("Documentos vigentes", documentos_vigentes, delta=f"Vencidos: {documentos_vencidos}")
        with kpi_col8:
            st.metric("Versiones pendientes", versiones_pendientes)

        serie_nc_registradas = construir_serie_mensual(df_nc_ind, "Fecha detección", "NC registradas", meses_analisis)
        serie_nc_cerradas = construir_serie_mensual(df_nc_ind[df_nc_ind["Estado"] == "Cerrada"], "Fecha cierre", "NC cerradas", meses_analisis) if not df_nc_ind.empty else pd.DataFrame()
        serie_acc_registradas = construir_serie_mensual(df_acc_ind, "Fecha inicio", "Acciones registradas", meses_analisis)
        serie_acc_cerradas = construir_serie_mensual(df_acc_ind[df_acc_ind["Estado"] == "Cerrada"], "Fecha cierre", "Acciones cerradas", meses_analisis) if not df_acc_ind.empty else pd.DataFrame()

        df_tendencia_nc = combinar_series_mensuales(serie_nc_registradas, serie_nc_cerradas)
        df_tendencia_acc = combinar_series_mensuales(serie_acc_registradas, serie_acc_cerradas)

        graf_col1, graf_col2 = st.columns(2)
        with graf_col1:
            st.write("### Tendencia mensual de no conformidades")
            if not df_tendencia_nc.empty:
                st.line_chart(df_tendencia_nc.set_index("Mes"))
            else:
                st.info("Todavía no hay datos suficientes para esta tendencia.")
        with graf_col2:
            st.write("### Tendencia mensual de acciones")
            if not df_tendencia_acc.empty:
                st.line_chart(df_tendencia_acc.set_index("Mes"))
            else:
                st.info("Todavía no hay datos suficientes para esta tendencia.")

        dist_col1, dist_col2 = st.columns(2)
        with dist_col1:
            st.write("### NC abiertas por área")
            if not df_nc_ind.empty:
                abiertas_por_area = df_nc_ind[df_nc_ind["Estado"] != "Cerrada"]["Área"].value_counts()
                if not abiertas_por_area.empty:
                    st.bar_chart(abiertas_por_area)
                else:
                    st.info("No hay no conformidades abiertas para este corte.")
            else:
                st.info("Sin datos de no conformidades.")
        with dist_col2:
            st.write("### Acciones vencidas por responsable")
            if not df_acc_ind.empty:
                vencidas_responsable = df_acc_ind[
                    (df_acc_ind["Estado"] != "Cerrada")
                    & (pd.to_datetime(df_acc_ind["Fecha compromiso"], errors="coerce") < pd.Timestamp.today().normalize())
                ]["Responsable"].value_counts()
                if not vencidas_responsable.empty:
                    st.bar_chart(vencidas_responsable)
                else:
                    st.info("No hay acciones vencidas registradas.")
            else:
                st.info("Sin datos de acciones.")

        dist_col3, dist_col4 = st.columns(2)
        with dist_col3:
            st.write("### Hallazgos por severidad")
            if not df_hall_ind.empty:
                st.bar_chart(df_hall_ind["Severidad"].value_counts())
            else:
                st.info("Sin hallazgos de auditoría registrados.")
        with dist_col4:
            st.write("### Estado documental")
            if not df_docs_ind.empty:
                st.bar_chart(df_docs_ind["Estado"].value_counts())
            else:
                st.info("Sin documentos registrados.")

        tabla_indicadores = pd.DataFrame(
            [
                {"Indicador": "Porcentaje de cierre de no conformidades", "Valor": formatear_porcentaje(nc_cerradas, total_nc), "Lectura": "Mide la capacidad de cerrar hallazgos detectados."},
                {"Indicador": "Porcentaje de cierre de acciones", "Valor": formatear_porcentaje(acciones_cerradas, total_acciones), "Lectura": "Mide el seguimiento real a acciones correctivas y preventivas."},
                {"Indicador": "Tiempo promedio de cierre de NC", "Valor": f"{promedio_cierre_nc} días", "Lectura": "Entre menor sea, más rápido responde el sistema de calidad."},
                {"Indicador": "Tiempo promedio de cierre de acciones", "Valor": f"{promedio_cierre_acciones} días", "Lectura": "Ayuda a detectar cuellos de botella en la ejecución."},
                {"Indicador": "Cumplimiento de auditorías", "Valor": formatear_porcentaje(auditorias_cerradas, total_auditorias), "Lectura": "Mide disciplina sobre auditorías planeadas."},
                {"Indicador": "Documentos vigentes sin vencer", "Valor": max(documentos_vigentes - documentos_vencidos, 0), "Lectura": "Refleja el nivel de control documental efectivo."},
            ]
        )

        st.write("### Resumen de indicadores")
        st.dataframe(tabla_indicadores, use_container_width=True, hide_index=True)

    with tab7:
        tarjeta_seccion(
            "Backup",
            "Respaldo y recuperación",
            "Protege base, evidencias y documentos con paquetes descargables listos para resguardo.",
        )

        if tiene_rol("calidad"):
            respaldos = listar_respaldos()
            ultimo_respaldo = respaldos[0] if respaldos else None

            info_col1, info_col2, info_col3 = st.columns(3)
            with info_col1:
                st.metric("Respaldos disponibles", len(respaldos))
            with info_col2:
                st.metric("Base actual", DB_NAME.name)
            with info_col3:
                st.metric(
                    "Último respaldo",
                    ultimo_respaldo.name if ultimo_respaldo else "Sin respaldo",
                )

            st.write(
                "Este módulo empaqueta `iner_voluntarios.db`, `evidencias_calidad` y `documentos_calidad` "
                "en un solo `.zip` dentro de la carpeta `backups`."
            )

            if st.button("Generar respaldo manual", use_container_width=True):
                try:
                    ruta_respaldo = crear_respaldo_sistema()
                    st.success(f"Respaldo generado correctamente: {ruta_respaldo.name}")
                    st.rerun()
                except Exception as e:
                    st.error(f"No se pudo generar el respaldo: {e}")

            respaldos = listar_respaldos()
            if respaldos:
                st.write("### Respaldos disponibles")
                tabla_respaldos = []
                for ruta in respaldos:
                    tabla_respaldos.append(
                        {
                            "Archivo": ruta.name,
                            "Fecha": datetime.fromtimestamp(ruta.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                            "Tamaño (KB)": round(ruta.stat().st_size / 1024, 1),
                        }
                    )
                st.dataframe(pd.DataFrame(tabla_respaldos), use_container_width=True, hide_index=True)

                st.write("### Descarga de respaldos")
                for indice, ruta in enumerate(respaldos[:10], start=1):
                    with open(ruta, "rb") as archivo_respaldo:
                        st.download_button(
                            label=f"Descargar respaldo {indice}: {ruta.name}",
                            data=archivo_respaldo.read(),
                            file_name=ruta.name,
                            mime="application/zip",
                            key=f"descarga_respaldo_{ruta.name}",
                        )
            else:
                st.info("Todavía no hay respaldos generados.")
        else:
            mostrar_aviso_permiso("solo Calidad o Admin pueden generar y descargar respaldos del sistema.")


st.sidebar.title("Navegación")
seccion = st.sidebar.radio(
    "Selecciona un módulo",
    ["C23-25", "B37-25", "Calidad"],
)

if seccion == "C23-25":
    mostrar_biobanco()
elif seccion == "B37-25":
    mostrar_proyecto_hospitalario()
elif seccion == "Calidad":
    mostrar_calidad()

st.sidebar.write(f"Sesión: {st.session_state.get('usuario_email', '')}")
st.sidebar.write(f"Perfil: {rol_actual().capitalize()}")
if st.session_state.get("es_admin", False):
    with st.sidebar:
        st.subheader("Aprobación de usuarios")

        pendientes = obtener_usuarios_pendientes()

        if pendientes:
            for id_usuario, email, fecha_registro in pendientes:
                st.write(f"{email} - registrado el {fecha_registro}")
                rol_aprobacion = st.selectbox(
                    f"Rol para {email}",
                    ["captura", "responsable", "auditor", "calidad", "admin"],
                    format_func=lambda valor: valor.capitalize(),
                    key=f"rol_aprobacion_{id_usuario}",
                )
                if st.button("Aprobar", key=f"aprobar_{id_usuario}", use_container_width=True):
                    aprobar_usuario(id_usuario, rol_aprobacion)
                    st.success(f"Usuario {email} aprobado.")
                    st.rerun()
        else:
            st.info("No hay usuarios pendientes.")

        st.subheader("Roles activos")
        usuarios_registrados = listar_usuarios()
        usuarios_aprobados = [fila for fila in usuarios_registrados if fila[2] == 1]
        if usuarios_aprobados:
            for id_usuario, email, _, _, rol, _ in usuarios_aprobados:
                nuevo_rol = st.selectbox(
                    email,
                    ROLES_USUARIO,
                    index=ROLES_USUARIO.index(rol if rol in ROLES_USUARIO else "captura"),
                    format_func=lambda valor: valor.capitalize(),
                    key=f"rol_usuario_{id_usuario}",
                )
                if st.button("Actualizar rol", key=f"actualizar_rol_{id_usuario}", use_container_width=True):
                    actualizar_rol_usuario(id_usuario, nuevo_rol)
                    st.success(f"Rol de {email} actualizado a {nuevo_rol}.")
                    st.rerun()
        else:
            st.info("Todavía no hay usuarios aprobados para administrar.")

if st.sidebar.button("Cerrar sesión"):
    st.session_state["autenticado"] = False
    st.session_state["usuario_email"] = ""
    st.session_state["es_admin"] = False
    st.session_state["rol_usuario"] = "captura"
    st.rerun()
